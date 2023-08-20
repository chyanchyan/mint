import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter


def fit_col_width(writer, df, sheet_name):
    # 自适应列宽
    column_widths = (
        df.columns.to_series().apply(lambda x: len(x.encode('utf-8'))).values
    )

    max_widths = (
        df.astype(str).applymap(lambda x: len(x.encode('utf-8'))).agg(max).values
    )

    max_widths = pd.Series(max_widths).fillna(0).to_list()
    widths = np.max([column_widths, max_widths], axis=0)

    worksheet = writer.sheets[sheet_name]
    for i, width in enumerate(widths, 1):
        if 'date' in df.columns[i - 1] or df.columns[i - 1][-1] == '日':
            worksheet.column_dimensions[get_column_letter(i)].width = 22
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = max(min(width + 2, 25), 10)



