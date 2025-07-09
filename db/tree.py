"""
数据库树形结构处理模块

该模块提供了用于处理数据库表之间关系的数据结构，支持：
1. 表之间的父子关系管理
2. 数据的层级查询和操作
3. 数据的导入导出（SQL、Excel）
4. JSON序列化

主要类：
- Tree: 基础树形结构类，处理表关系
- DataTree: 带数据的树形结构类，继承自Tree
"""

import pandas as pd
from pandas import ExcelWriter
from openpyxl import Workbook
from copy import copy
import os
import sys

# 添加父目录到系统路径，以便导入mint模块
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)

if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# 导入辅助函数
from mint.helper_function.hf_xl import fit_col_width
from mint.helper_function.hf_func import *
from mint.helper_function.hf_array import get_crop_from_df
from mint.helper_function.hf_data import *


# @profile_line_by_line
def get_cst_pki(con, schemas):
    """
    获取数据库约束信息
    
    Args:
        con: 数据库连接对象
        schemas: 数据库模式列表
    
    Returns:
        pandas.DataFrame: 包含约束信息的DataFrame
    """
    in_str = ', '.join([f'"{schema}"' for schema in schemas])
    stmt = (
        "select * from information_schema.KEY_COLUMN_USAGE "
        f"where CONSTRAINT_SCHEMA in ({in_str})"
    )

    res = pd.read_sql(sql=stmt, con=con)
    return res


# @profile_line_by_line
def get_booking_sequence(cst_pki, root_nodes=None):
    """
    获取表的录入顺序（拓扑排序）
    
    Args:
        cst_pki: 约束信息DataFrame
        root_nodes: 根节点列表，如果指定则只处理相关表
    
    Returns:
        list: 表的录入顺序
    """
    relation_info = cst_pki[['TABLE_NAME', 'REFERENCED_TABLE_NAME']].values.tolist()
    if root_nodes is not None:
        related_tables = set()
        graph = get_graph(relation_info=relation_info)
        for node in root_nodes:
            r = get_related_nodes(
                graph=graph,
                node=node
            )
            related_tables |= set(r)
        relation_info = cst_pki[
            cst_pki['TABLE_NAME'].isin(related_tables) |
            cst_pki['REFERENCED_TABLE_NAME'].isin(related_tables)
        ].replace(np.nan, None)[['TABLE_NAME', 'REFERENCED_TABLE_NAME']].values.tolist()

    booking_seq = topological_sort(relation_info=relation_info)

    return booking_seq


class Tree(JsonObj):
    """
    基础树形结构类
    
    用于表示数据库表之间的层级关系，支持父子关系的管理
    和拓扑排序等功能。
    """
    
    def __init__(
            self,
            con,
            tables,
            tree=None,
            root=None,
            ref=None,
            reffed=None,
            is_parent=False,
            parent_root=None,
    ):
        """
        初始化树形结构
        
        Args:
            con: 数据库连接对象
            tables: 表对象字典
            tree: 现有树对象（用于复制）
            root: 根表名
            ref: 引用列名
            reffed: 被引用列名
            is_parent: 是否为父节点
            parent_root: 父根节点名
        """
        if tree:
            # 如果提供了现有树对象，直接复制其属性
            self.con = tree.con
            self.root = tree.root
            self.ref = tree.ref
            self.reffed = tree.reffed
            self.table = tree.table
            self.tables = tree.tables
            self.parents = tree.parents
            self.children = tree.children
            self.node_names = tree.node_names
            self.booking_sequence = tree.booking_sequence
            self.reading_sequence = tree.reading_sequence
            return

        # 初始化基本属性
        self.con = con
        self.root = root
        self.schemas = list(set([table.schema for table in tables.values()]))
        self.tables = tables
        self.table = tables[root]
        self.ref = ref
        self.reffed = reffed
        self.parents = []
        self.children = []

        JsonObj.__init__(self)

        # 处理父节点关系
        ref_cols = [
            col for col in self.table.cols if not pd.isna(col.foreign_key)
        ]

        for ref_col in ref_cols:
            fk = ref_col.foreign_key
            col_parent_root = fk.split('.')[1]
            if col_parent_root == parent_root:
                continue
            ref = ref_col.col_name
            reffed = fk.split('.')[2]

            parent = Tree(
                con=con,
                tables=self.tables,
                root=col_parent_root,
                ref=ref,
                reffed=reffed,
                is_parent=True
            )
            self.parents.append(parent)

        # 处理子节点关系
        if not is_parent:
            reffed_table_and_cols = []
            for child_root, table_obj in tables.items():
                if child_root != self.root:
                    for col_obj in table_obj.cols:
                        fk = col_obj.foreign_key
                        if not pd.isna(fk):
                            schema, reffed_table_name, reffed = fk.split('.')
                            if reffed_table_name == self.root:
                                reffed_table_and_cols.append((child_root, col_obj.col_name, reffed))
            for child_root, ref, reffed in reffed_table_and_cols:
                child = Tree(
                    con=con,
                    tables=self.tables,
                    root=child_root,
                    ref=ref,
                    reffed=reffed,
                    parent_root=self.root
                )
                self.children.append(child)

        # 按表顺序排序父节点和子节点
        self.parents.sort(key=lambda x: self.tables[x.root].order)
        self.children.sort(key=lambda x: self.tables[x.root].order)

        # 生成录入顺序和读取序列
        self.booking_sequence = self.get_booking_sequence()
        self.reading_sequence = [item for item in self.booking_sequence.__reversed__()]

        self.node_names = self.get_node_names()
        # self.sort_cols()

    def __bool__(self):
        """判断树是否有效"""
        if self.root:
            return True
        else:
            return False

    def __gt__(self, other):
        """判断当前节点是否为其他节点的祖先"""
        return self.root in [p for p in other.all_parenthood_names()]

    def __lt__(self, other):
        """判断当前节点是否为其他节点的后代"""
        return self.root in [c for c in other.all_childhood_names()]

    def __ge__(self, other):
        """判断当前节点是否为其他节点的祖先或等于其他节点"""
        return self.root in [p for p in other.all_parenthood_names()] or self.root == other.root

    def __le__(self, other):
        """判断当前节点是否为其他节点的后代或等于其他节点"""
        return self.root in [c for c in other.all_childhood_names()] or self.root == other.root

    def __eq__(self, other):
        """判断两个节点是否相等"""
        return self.root == other.root

    def get_booking_sequence(self):
        """
        获取录入顺序（拓扑排序）
        
        Returns:
            list: 包含所有相关表的录入顺序
        """
        res = sum([
            p.get_booking_sequence() for p in self.parents
        ], start=[]) + [
            self.root
        ] + sum([
            c.get_booking_sequence() for c in self.children
        ], start=[])

        return [item for i, item in enumerate(res) if item not in res[:i]]

    def all_parenthood_names(self):
        """
        获取所有祖先节点名称
        
        Returns:
            set: 祖先节点名称集合
        """
        res = set([p.root for p in self.parents])
        for p in self.parents:
            res |= p.all_parenthood_names()
        for c in self.children:
            res |= c.all_parenthood_names()
        return res

    def all_childhood_names(self):
        """
        获取所有后代节点名称
        
        Returns:
            set: 后代节点名称集合
        """
        res = set([c.root for c in self.children])
        for c in self.children:
            res |= c.all_childhood_names()
        return res

    def get_child_path(self, child_root: str):
        """
        获取到指定子节点的路径
        
        Args:
            child_root: 子节点名称
            
        Returns:
            list: 从当前节点到子节点的路径
        """
        res = []
        for c in self.children:
            if c.root == child_root:
                return [self.root, c.root]

            res.extend(c.get_child_path(child_root=child_root))
        if len(res) > 0:
            res = [self.root] + res

        return res

    def json_obj_base(self):
        """
        获取基础JSON对象表示
        
        Returns:
            dict: 树的JSON表示
        """
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'ref', 'reffed', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        for parent in self.parents:
            json_obj['parents'].append(parent.json_obj_base())
        for child in self.children:
            json_obj['children'].append(child.json_obj_base())
        json_obj['table'] = self.table.to_json_obj()
        return json_obj

    @property
    def json_obj(self):
        """获取JSON对象表示"""
        return self.json_obj_base()

    def p_(self, p_name, ref=None):
        """
        获取指定父节点
        
        Args:
            p_name: 父节点名称
            ref: 引用列名（可选）
            
        Returns:
            Tree: 父节点对象，如果不存在则返回None
        """
        if not ref:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name
                ), None
            )
        else:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name and parent.ref == ref
                ), None
            )

    def c_(self, c_name, reffed=None):
        """
        获取指定子节点
        
        Args:
            c_name: 子节点名称
            reffed: 被引用列名（可选）
            
        Returns:
            Tree: 子节点对象，如果不存在则返回None
        """
        if not reffed:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name
                ), None
            )
        else:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name and child.reffed == reffed
                ), None
            )

    @property
    def p(self):
        """父节点属性索引器"""
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        """子节点属性索引器"""
        return PropertyIndexer(self, 'c_')

    def get_all_nodes(self, res=None):
        """
        获取所有节点
        
        Args:
            res: 结果字典（递归使用）
            
        Returns:
            dict: 所有节点的字典，键为节点名，值为节点对象
        """
        if not res:
            res = {self.root: self}
        for p in self.parents:
            if p.root not in list(res.keys()):
                res[p.root] = p
                res.update(p.get_all_nodes(res=res))
        for c in self.children:
            if c.root not in list(res.keys()):
                res[c.root] = c
            res.update(c.get_all_nodes(res=res))
        return res

    def get_all_parents(self, res=None):
        """
        获取所有父节点
        
        Args:
            res: 结果字典（递归使用）
            
        Returns:
            dict: 所有父节点的字典
        """
        if not res:
            res = {}
        for p in self.parents:
            if p.root not in list(res.keys()):
                res[p.root] = p
                res.update(p.get_all_parents(res=res))
        for c in self.children:
            res.update(c.get_all_parents(res=res))
        return res

    def get_node_names(self):
        """
        获取所有节点名称
        
        Returns:
            set: 节点名称集合
        """
        res = {self.root}
        for parent in self.parents:
            res |= parent.node_names
        for child in self.children:
            res |= child.node_names
        return res

    def get_parent_tree_structure(self):
        """
        获取父树结构
        
        Returns:
            dict: 父树结构字典
        """
        res = {p.root: p.get_parent_tree_structure() for p in self.parents}
        return res


class DataTree(Tree):
    """
    带数据的树形结构类
    
    继承自Tree类，增加了数据操作功能，支持数据的查询、导入、导出等操作。
    """
    
    def __init__(
            self,
            con=None,
            tables=None,
            tree=None,
            root=None,
            ref=None,
            reffed=None,
            relevant_data_set=None,
    ):
        """
        初始化数据树
        
        Args:
            con: 数据库连接对象
            tables: 表对象字典
            tree: 现有树对象
            root: 根表名
            ref: 引用列名
            reffed: 被引用列名
            relevant_data_set: 相关数据集
        """
        if con is None:
            if tree is None:
                print('con and tree cannot be None at the same time')
                raise ValueError
            else:
                con = copy(tree.con)
        if tables is None:
            if tree is None:
                print('tables and tree cannot be None at the same time')
                raise ValueError
            else:
                tables = copy(tree.tables)

        self.schemas = [table.schema for table in tables.values()]

        super().__init__(
            con=con,
            tree=tree,
            root=root,
            tables=tables,
            ref=ref,
            reffed=reffed,
        )

        # 初始化数据
        if relevant_data_set is None:
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=[col.col_name for col in self.tables[node_name].cols])
                 for node_name in self.node_names]
            ))
        try:
            self.data = relevant_data_set[self.root]
        except KeyError:
            self.data = pd.DataFrame(columns=[col.col_name for col in self.table.cols])
        self.relevant_data_set = relevant_data_set

    def __len__(self):
        """返回数据行数"""
        return len(self.data)

    def __getitem__(self, key):
        """获取数据列"""
        return self.data[key]

    @property
    def loc(self):
        """DataFrame的loc访问器"""
        return self.data.loc

    @property
    def iloc(self):
        """DataFrame的iloc访问器"""
        return self.data.iloc

    def iterrows(self):
        """迭代数据行"""
        return self.data.iterrows()

    def itertrees(self):
        """
        迭代树结构
        
        Yields:
            tuple: (索引, DataTree对象)
        """
        for i, row in self.iterrows():
            relevant_data_set = deepcopy(self.relevant_data_set)
            relevant_data_set[self.root] = row.to_frame().transpose()
            datatree = DataTree(
                con=self.con,
                tables=self.tables,
                tree=self,
                relevant_data_set={}
            )
            datatree.from_relevant_data_set(relevant_data_set)
            yield i, datatree

    @staticmethod
    def update_values_map(values_map, tree, root_data):
        """
        更新值映射
        
        Args:
            values_map: 值映射字典
            tree: 树对象
            root_data: 根数据
            
        Returns:
            dict: 更新后的值映射
        """
        for p in tree.parents:
            if len(root_data) == 0:
                adding_data = set()
            else:
                try:
                    adding_data = set(root_data[p.ref].dropna())
                except KeyError:
                    print(f'parent ref info error: {p.ref}')
                    adding_data = set()

            try:
                values_map[p.root][p.reffed]['values'] = \
                    values_map[p.root][p.reffed]['values'] | adding_data
            except KeyError:
                values_map[p.root] = {}
                values_map[p.root][p.reffed] = {
                    'tree': p,
                    'values': adding_data
                }

        for c in tree.children:
            if len(root_data) == 0:
                adding_data = set()
            else:
                adding_data = set(root_data[c.reffed].dropna())

            try:
                values_map[c.root][c.ref]['values'] = \
                    values_map[c.root][c.ref]['values'] | adding_data
            except KeyError:
                values_map[c.root] = {}
                values_map[c.root][c.ref] = {
                    'tree': c,
                    'values': adding_data
                }

        return values_map

    def fill_na_with_none(self):
        """将NaN值替换为None"""
        relevant_data_set_copy = deepcopy(self.relevant_data_set)
        for name, df in relevant_data_set_copy.items():
            df = df.fillna(np.nan).replace([np.nan], [None])
            self.relevant_data_set[name] = df

    def from_sql(
            self,
            index_col: str = None,
            index_values: Set[str] = None,
            limit=None,
            offset=None
    ):
        """
        从SQL查询加载数据
        
        Args:
            index_col: 索引列名
            index_values: 索引值集合
            limit: 限制行数
            offset: 偏移量
        """
        # 查询根数据
        root_schema = self.table.schema
        if index_col is None and index_values is None:
            where_str = ''
        else:
            if index_col is None:
                index_col = self.table.pk
            if index_values is not None and len(index_values) == 0:
                limit = 0
            values_str = ', '.join([f'"{index_value}"' for index_value in index_values])
            where_str = f'WHERE `{index_col}` in ({values_str})'

        if limit is None:
            limit_str = ''
        else:
            limit_str = f'limit {str(limit)}'

        if offset is None:
            offset_str = ''
        else:
            offset_str = f'offset {str(offset)}'

        sql = f'SELECT * FROM {root_schema}.{self.root} {where_str} {limit_str} {offset_str}'
        root_data = pd.read_sql(sql=sql, con=self.con, index_col=self.table.pk)
        if len(root_data) == 0:
            print(f'empty result for root: "{self.root}" \n'
                  f'with index col: "{where_str}" \n'
                  f'with values in {index_values}')
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=[col.col_name for col in self.tables[node_name].cols])
                 for node_name in self.node_names]
            ))
        else:
            relevant_data_set = {self.root: root_data}

            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # 查询相关数据
            reading_seq = list(self.reading_sequence)
            reading_seq.remove(self.root)

            for node_name in reading_seq:
                where_str_list = []
                if node_name not in values_map:
                    continue
                for col, tree_values in values_map[node_name].items():
                    values = tree_values['values']

                    # 该列没有值
                    if len(values) == 0:
                        continue

                    values_str = ', '.join(['"%s"' % value.replace('%', '%%') for value in values])
                    where_str = f'(`{col}` in ({values_str}))'
                    where_str_list.append(where_str)

                # 该表没有值
                if len(where_str_list) == 0:
                    continue

                all_where_str = ' or '.join(where_str_list)

                sql = f'SELECT * FROM {self.tables[node_name].schema}.{node_name} WHERE {all_where_str}'
                data = pd.read_sql(sql=sql, con=self.con, index_col='id')
                relevant_data_set[node_name] = data

                for col, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

            for node_name, value_map in values_map.items():
                if node_name not in relevant_data_set:
                    where_str_list = []
                    for col, tree_values in values_map[node_name].items():
                        values = tree_values['values']

                        # 该列没有值
                        if len(values) == 0:
                            continue

                        values_str = ', '.join(['"%s"' % value.replace('%', '%%') for value in values])
                        where_str = f'(`{col}` in ({values_str}))'
                        where_str_list.append(where_str)

                    # 该表没有值
                    if len(where_str_list) == 0:
                        continue
                    all_where_str = ' or '.join(where_str_list)

                    sql = f'SELECT * FROM {self.tables[node_name].schema}.{node_name} WHERE {all_where_str}'
                    data = pd.read_sql(sql=sql, con=self.con, index_col='id')
                    relevant_data_set[node_name] = data

        self.data = relevant_data_set[self.root]
        self.relevant_data_set = relevant_data_set

    def from_relevant_data_set(self, relevant_data_set=None):
        """
        从相关数据集加载数据
        
        Args:
            relevant_data_set: 相关数据集字典
        """
        relevant_data_set_copy = copy(relevant_data_set)

        if relevant_data_set_copy is None:
            relevant_data_set_copy = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=[col.col_name for col in self.tables[node_name].cols])
                 for node_name in self.node_names]
            ))

        try:
            root_data = relevant_data_set_copy[self.root]
        except KeyError:
            root_data = pd.DataFrame(columns=[col.col_name for col in self.table.cols])

        if len(root_data) == 0:
            relevant_data_set_copy = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=[col.col_name for col in self.tables[node_name].cols])
                 for node_name in self.node_names]
            ))
        else:
            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # 处理相关数据
            for node_name in self.reading_sequence:
                if node_name not in relevant_data_set_copy:
                    relevant_data_set_copy[node_name] = pd.DataFrame(
                        columns=[col.col_name for col in self.tables[node_name].cols]
                    )
                    continue
                if node_name not in values_map:
                    continue
                where_str_list = []
                data = pd.DataFrame(columns=[col.col_name for col in self.table.cols])
                for col, tree_values in values_map[node_name].items():
                    values = tree_values['values']
                    values_str = ', '.join([f'"{value}"' for value in values])
                    data = relevant_data_set_copy[node_name]
                    where_str = f'(data["{col}"].isin([{values_str}]))'
                    where_str_list.append(where_str)

                all_where_str = ' | '.join(where_str_list)

                try:
                    relevant_data_set_copy[node_name] = eval(f'data[{all_where_str}]')
                except SyntaxError:
                    print(where_str_list)
                    print(all_where_str)
                    raise SyntaxError

                for col, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

        self.data = root_data
        self.relevant_data_set = relevant_data_set_copy

    @check_param_valid_range(
        [
            'if_data_exists'
        ],
        [
            ['fail', 'skip', 'merge', 'update']
        ]
    )
    def to_sql(self, curd, if_data_exists='fail'):
        """
        将数据保存到SQL数据库
        
        Args:
            curd: CRUD操作对象
            if_data_exists: 数据存在时的处理策略
        """
        stmt = curd.stmt_df_merge(
            datatree=self,
            if_data_exists=if_data_exists,
        )
        print(stmt)
        # con.cursor.execute(stmt)

    def p_(self, p_name, ref=None):
        """
        获取指定父节点（带数据）
        
        Args:
            p_name: 父节点名称
            ref: 引用列名（可选）
            
        Returns:
            DataTree: 父节点数据树对象
        """
        p_tree = Tree.p_(self, p_name=p_name, ref=ref)
        p = DataTree(
            con=self.con,
            tables=self.tables,
            tree=p_tree
        )
        p.from_relevant_data_set(self.relevant_data_set)
        return p

    def c_(self, c_name, reffed=None):
        """
        获取指定子节点（带数据）
        
        Args:
            c_name: 子节点名称
            reffed: 被引用列名（可选）
            
        Returns:
            DataTree: 子节点数据树对象
        """
        c = DataTree(
            con=self.con,
            tables=self.tables,
            tree=Tree.c_(self, c_name=c_name, reffed=reffed)
        )
        c.from_relevant_data_set(self.relevant_data_set)
        return c

    @property
    def p(self):
        """父节点属性索引器（带数据）"""
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        """子节点属性索引器（带数据）"""
        return PropertyIndexer(self, 'c_')

    @property
    def ps(self):
        """
        父节点生成器
        
        Yields:
            DataTree: 父节点数据树对象
        """
        for parent in self.parents:
            p = DataTree(
                con=self.con,
                tables=self.tables,
                tree=parent
            )
            p.from_relevant_data_set(self.relevant_data_set)
            yield p

    @property
    def cs(self):
        """
        子节点生成器
        
        Yields:
            DataTree: 子节点数据树对象
        """
        for child in self.children:
            c = DataTree(
                con=self.con,
                tables=self.tables,
                tree=child
            )
            c.from_relevant_data_set(self.relevant_data_set)
            yield c

    def json_obj_base(self, with_value=False, with_table=True):
        """
        获取基础JSON对象表示（带数据）
        
        Args:
            with_value: 是否包含数据值
            with_table: 是否包含表信息
            
        Returns:
            dict: 数据树的JSON表示
        """
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'ref', 'reffed', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        data = self.data.copy(deep=True)
        data = data.replace(np.nan, None)
        for parent in self.ps:
            json_obj['parents'].append(parent.json_obj_base(with_value=with_value, with_table=with_table))
        for child in self.cs:
            json_obj['children'].append(child.json_obj_base(with_value=with_value, with_table=with_table))

        if with_value:
            json_obj['values'] = data.to_dict(orient='records')

        if with_table:
            json_obj['table'] = self.table.to_json_obj()

        return json_obj

    @property
    def json_obj(self):
        """获取完整JSON对象表示"""
        return self.json_obj_base(with_value=True, with_table=True)

    def json_obj_func(self, with_value=True, with_table=True):
        """
        获取JSON对象表示（可配置）
        
        Args:
            with_value: 是否包含数据值
            with_table: 是否包含表信息
            
        Returns:
            dict: 数据树的JSON表示
        """
        return self.json_obj_base(with_value=with_value, with_table=with_table)

    # @profile_line_by_line
    def nested_values(self, ref_group=None, ignore_ref_col=None, full_detail=False):
        """
        获取嵌套值结构
        
        Args:
            ref_group: 引用分组
            ignore_ref_col: 忽略的引用列
            full_detail: 是否包含完整详情
            
        Returns:
            dict: 嵌套值结构
        """
        if full_detail:
            col_items = [
                item for item in self.table.cols
                if not item[1].col_name == ignore_ref_col
            ]
        else:
            col_items = [
                item for item in self.table.cols
                if not pd.isna(item[1].web_list_order)
                and not item[1].col_name == ignore_ref_col
            ]
        col_items.sort(key=lambda x: x[1].web_list_order)
        display_column_names, display_columns = zip(*col_items)
        display_column_names = ['key'] + list(display_column_names)
        columns = [
            {
                'key': i,
                'dataIndex': col.col_name,
                'title': col.label,
                'dataType': col.data_type
            } for i, col in enumerate(display_columns)
        ]
        data = self.data.assign(key=self.data.index.astype(str))
        data['id'] = data.index
        data.sort_index()
        display_data = data[display_column_names]

        if not ref_group:
            data_source = {0: display_data.to_dict(orient='records')}
        else:
            data_source = {
                key: display_data.loc[names].to_dict(orient='records')
                for key, names in ref_group.items()
            }

        children = {}
        for child in self.cs:
            child_data = child.data.sort_index()
            gs_data = {key: value for key, value in data.groupby(by=child.reffed)}
            ref_group = {
                str(gs_data[ref_value].index[0]): group.index.to_list()
                for ref_value, group in child_data.groupby(by=child.ref)
            }
            children[child.root] = child.nested_values(
                ref_group=ref_group,
                ignore_ref_col=child.ref,
                full_detail=full_detail
            )

        res = {
            'tag': self.table.label,
            'columns': columns,
            'dataSource': data_source,
            'children': children
        }

        return res

    # @profile_line_by_line
    def nested_values2(self, data=None, parent_ref=None):
        """
        获取嵌套值结构（版本2）
        
        Args:
            data: 数据DataFrame
            parent_ref: 父引用列
            
        Returns:
            dict: 嵌套值结构
        """
        if not isinstance(data, pd.DataFrame):
            data = self.data
        col_items = [
            item for item in self.table.cols
            if not pd.isna(item[1].web_list_order)
            and item[1].col_name != parent_ref
        ]
        col_items.sort(key=lambda x: x[1].web_list_order)
        display_column_names, display_columns = zip(*col_items)
        display_column_names = ['key'] + list(display_column_names)
        columns = [
            {
                'key': i,
                'dataIndex': col.col_name,
                'title': col.label,
                'dataType': col.data_type
            } for i, col in enumerate(display_columns)
        ]
        data = data.assign(key=data.index.astype(str))
        data['id'] = data.index
        data.sort_index()

        if not parent_ref:
            display_data = data[display_column_names]
            data_source = {0: display_data.to_dict(orient='records')}
        else:
            gs_data = data.groupby(by=parent_ref)
            data_source = {
                key: g_data[display_column_names].to_dict(orient='records')
                for key, g_data in gs_data
            }

        children = {}
        for child in self.cs:
            parent_data = pd.DataFrame(
                data=data[['id', child.reffed]]
            )
            parent_data.columns = [f'id_from_{self.root}', child.reffed]
            child_data = pd.merge(left=child.data,
                                  right=parent_data,
                                  left_on=child.ref, right_on=child.reffed,
                                  suffixes=('', f'_from_{self.root}'))
            children[child.root] = child.nested_values2(
                data=child_data,
                parent_ref=f'id_from_{self.root}'
            )

        res = {
            'tag': self.table.label,
            'columns': columns,
            'dataSource': data_source,
            'children': children
        }

        return res

    def get_all_parents_with_values(self, res=None):
        """
        获取所有带值的父节点
        
        Args:
            res: 结果字典（递归使用）
            
        Returns:
            dict: 所有带值的父节点字典
        """
        if res is None:
            res = {}
        for dp in self.ps:
            if dp.root not in res:
                res[dp.root] = dp
            # else:
            #     res[dp.root].data = pd.concat([res[dp.root].data, dp.data]).drop_duplicates(dp.reffed)
            res = dp.get_all_parents_with_values(res=res)

        for dc in self.cs:
            res = dc.get_all_parents_with_values(res=res)

        return res


    def get_all_parents_with_full_value(self):
        """
        获取所有带完整值的父节点
        
        Returns:
            dict: 所有带完整值的父节点字典
        """
        res = {}
        all_parents = self.get_all_parents()
        for p_name, p in all_parents.items():
            dp = DataTree(
                con=self.con,
                tables=self.tables,
                tree=p
            )
            dp.from_sql()
            res[p.root] = dp
        return res

    def get_parents_select_values(self):
        """
        获取父节点的选择值
        
        Returns:
            dict: 父节点选择值字典
        """
        res = {}
        all_parents = self.get_all_parents_with_full_value()
        for p_name, p in all_parents.items():
            try:
                res[p.root] = p.data[p.reffed].to_list()
            except KeyError:
                if p.data.index.name == p.reffed:
                    res[p.root] = p.data.index.to_list()
                else:
                    raise KeyError

        return res

    def from_excel_booking_sheet(self, dfs):
        """
        从Excel预订表加载数据
        
        Args:
            dfs: Excel工作表字典
        """
        relevant_data_set = {}
        root_sheet_name = f'bks_{self.table.label}'
        root_df_raw = dfs[root_sheet_name]
        root_df = get_crop_from_df(
            df=root_df_raw,
            anchor_x=3,
            anchor_y=4,
            vertical=True,
            col_offset=-2,
            pk_index=None
        )

        relevant_data_set[self.root] = root_df

        # 处理子节点数据
        for child in self.children:
            try:
                anchor_x = root_df_raw.iloc[:, 2].to_list().index(child.root) + 4
            except ValueError:
                print(f'child {child.root} not found in {root_sheet_name}')
                continue
            child_df = get_crop_from_df(
                df=root_df_raw,
                anchor_x=anchor_x,
                anchor_y=3,
                vertical=False,
                col_offset=-2,
                pk_index=-3
            )
            # 标记子节点引用
            child_df[child.ref] = root_df[child.reffed].values[0]

            # 标记子节点自动命名
            naming_cols = sorted([
                col_obj for col_obj in child.table.cols
                if not pd.isna(col_obj.naming_field_order)
            ], key=lambda x: x.naming_field_order)
            if len(naming_cols) > 0 and len(child_df) > 0:
                child_df['name'] = child_df.apply(
                    lambda x: '-'.join([
                        x[col_obj.col_name] for col_obj in naming_cols
                    ]),
                    axis=1
                )

            relevant_data_set[child.root] = child_df

        # 处理父节点数据
        for parent_root, parent in self.get_all_parents().items():
            parent_sheet_name = f'bks_{parent.table.label}'
            try:
                parent_df_raw = dfs[parent_sheet_name]
            except KeyError:
                print(f'no sheet for root "{parent_root}"')
                continue
            parent_df = get_crop_from_df(
                df=parent_df_raw,
                anchor_x=7,
                anchor_y=3,
                vertical=False,
                col_offset=-2,
                pk_index=-3
            )

            # 标记自动命名
            naming_cols = [
                (col.col_name, col.naming_field_order)
                for col in parent.table.cols
                if not pd.isna(col.naming_field_order)
            ]
            if len(naming_cols) > 0:
                naming_cols = [col for col, order in sorted(naming_cols, key=lambda x: x[1])]

                if len(parent_df) > 0:
                    if 'name' in parent_df.columns:
                        parent_df['name'] = parent_df.apply(
                            lambda x: '-'.join([x[col] for col in naming_cols]) if pd.isna(x['name']) else x[
                                'name'],
                            axis=1
                        )
                    else:
                        parent_df['name'] = parent_df.apply(
                            lambda x: '-'.join([x[col] for col in naming_cols]),
                            axis=1
                        )

            relevant_data_set[parent_root] = parent_df

        # 处理空值
        for key, df in relevant_data_set.items():
            relevant_data_set[key] = df.where(df.notnull(), None)

        self.from_relevant_data_set(relevant_data_set=relevant_data_set)

    def to_excel_sheets(self, pth, node_names=None, web_visible_only=False, label=False, table_label=False):
        """
        将数据导出到Excel工作表
        
        Args:
            pth: 文件路径
            node_names: 节点名称列表
            web_visible_only: 是否只导出web可见列
            label: 是否使用标签名
            table_label: 是否使用表标签
        """
        writer = ExcelWriter(pth, mode='w', engine='openpyxl', options={'strings_to_urls': False})
        if not node_names:
            node_names = self.node_names
        flag = False

        for node_name in node_names:
            table = self.tables[node_name]
            if web_visible_only:
                columns = [
                    col.col_name
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            else:
                columns = [
                    col.col_name
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            try:
                df = self.relevant_data_set[node_name][columns]
            except KeyError:
                df = pd.DataFrame(columns=columns)

            if label:
                column_mapper = dict(
                    [
                        (col.col_name, col.label)
                        for col in table.cols.values()]
                )
                df.rename(columns=column_mapper, inplace=True)
            if table_label:
                node_name = table.label

            df.to_excel(writer, sheet_name=node_name, index=False,
                        columns=df.columns)
            fit_col_width(writer=writer, df=df, sheet_name=node_name)
            flag = True

        if flag:
            writer._save()
            writer.close()
        else:
            del writer
            os.remove(pth)

    def to_excel_detail_sheet(self, pth):
        """
        将数据导出到Excel详细工作表
        
        Args:
            pth: 文件路径
        """
        child_data = {}
        for child in self.children:
            child_data[child.root] = self[child.root].to_dict()
        wb = Workbook()
        ws = wb.worksheets[0]
        pin = 1
        for r, (key, value) in enumerate(self.data.items()):
            if self.table.cols[key].web_visible == 1:
                ws.cell(row=pin, column=1).value = self.table.cols[key].label
                ws.cell(row=pin, column=2).value = value
                pin += 1

        pin += 1

        for child_name, child_values in child_data.items():
            ws.cell(row=pin, column=1).value = self.tables[child_name].comment
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    ws.cell(row=pin, column=c_index).value = self.tables[child_name].cols[col].label
                    c_index += 1
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    for r, (index, v) in enumerate(value.items()):
                        ws.cell(row=pin + r, column=c_index).value = v
                    c_index += 1
            pin += len(child_values) + 1

        wb.save(pth)


def get_right_angle_trees_from_tree(tree: Tree, res=None, include_root=True):
    """
    从树中获取直角树列表
    
    Args:
        tree: 树对象
        res: 结果列表（递归使用）
        include_root: 是否包含根节点
        
    Returns:
        list: 直角树列表
    """
    if include_root:
        if res is None:
            res = [tree]
        else:
            for item in res:
                if item.root == tree.root:
                    break
            else:
                res.append(tree)

    if tree.parents:
        for parent in tree.parents:
            get_right_angle_trees_from_tree(parent, res)
    if tree.children:
        for child in tree.children:
            get_right_angle_trees_from_tree(child, res, include_root=False)
    return res


def get_flattened_tree_list_from_right_angle_trees(right_angle_trees: List[Tree], res=None):
    """
    从直角树列表获取扁平化树列表
    
    Args:
        right_angle_trees: 直角树列表
        res: 结果列表（递归使用）
        
    Returns:
        list: 扁平化树列表
    """
    if res is None:
        res = []
    for tree in right_angle_trees:
        res.append(tree)
        get_flattened_tree_list_from_right_angle_trees(tree.children, res)
    return res


if __name__ == '__main__':
    pass