from pandas import ExcelWriter
from openpyxl import Workbook
from copy import copy
import os
import sys
from collections import OrderedDict

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)

if parent_dir not in sys.path:
    sys.path.append(parent_dir)

from mint.helper_function.hf_xl import fit_col_width
from mint.helper_function.hf_func import *
from mint.helper_function.hf_array import get_crop_from_df
from mint.helper_function.hf_data import *


# @profile_line_by_line
def get_cst_pki(con, schemas):
    in_str = ', '.join([f'"{schema}"' for schema in schemas])
    stmt = (
        "select * from information_schema.KEY_COLUMN_USAGE "
        f"where CONSTRAINT_SCHEMA in ({in_str})"
    )

    res = pd.read_sql(sql=stmt, con=con)
    return res


# @profile_line_by_line
def get_booking_sequence(cst_pki, root_nodes=None):
    relation_info = cst_pki[['TABLE_NAME', 'REFERENCED_TABLE_NAME']].values.tolist()
    if root_nodes is not None:
        related_tables = set()
        graph = get_graph(relation_info=relation_info)
        for node in root_nodes:
            r = get_related_nodes(
                graph=graph,
                node=node
            )
            related_tables |= set(r)
        relation_info = cst_pki[
            cst_pki['TABLE_NAME'].isin(related_tables) |
            cst_pki['REFERENCED_TABLE_NAME'].isin(related_tables)
        ].replace(np.nan, None)[['TABLE_NAME', 'REFERENCED_TABLE_NAME']].values.tolist()

    booking_seq = topological_sort(relation_info=relation_info)

    return booking_seq


class Tree(JsonObj):
    @profile_line_by_line
    def __init__(
            self,
            con,
            tables,
            tree=None,
            root=None,
            cst=None,
            ref=None,
            reffed=None
    ):
        if tree:
            self.con = tree.con
            self.tables = tree.tables
            self.root = tree.root
            self.cst = copy(tree.cst)
            self.ref = tree.ref
            self.reffed = tree.reffed
            self.table = tree.table

            self.ref_info = tree.cst[
                (tree.cst['CONSTRAINT_NAME'] != 'PRIMARY') &
                (tree.cst['TABLE_NAME'] == tree.root) &
                ~(pd.isna(tree.cst['REFERENCED_TABLE_NAME']))
            ]
            self.reffed_info = tree.cst[
                tree.cst['REFERENCED_TABLE_NAME'] == tree.root
            ]

            self.parents = tree.parents
            self.children = tree.children
            self.node_names = tree.node_names
            self.booking_sequence = tree.booking_sequence
            self.reading_sequence = tree.reading_sequence
            return

        self.con = con
        self.root = root

        self.schemas = list(set([table.schema for table in tables.values()]))
        if cst is None:
            cst_pki = get_cst_pki(con=con, schemas=self.schemas)
            cst = cst_pki[cst_pki['CONSTRAINT_NAME'] != 'PRIMARY']
        self.cst = deepcopy(cst)

        self.tables = tables
        self.table = tables[root]

        self.ref = ref
        self.reffed = reffed

        self.parents = []
        self.children = []

        JsonObj.__init__(self)

        self.ref_info = cst[
            (cst['TABLE_NAME'] == root) &
            ~(pd.isna(cst['REFERENCED_TABLE_NAME']))
        ]
        self.reffed_info = cst[cst['REFERENCED_TABLE_NAME'] == root]

        for i, parents_cst_row in self.ref_info.iterrows():
            parent_root = parents_cst_row['REFERENCED_TABLE_NAME']
            parent_ref = parents_cst_row['COLUMN_NAME']
            parent_reffed = parents_cst_row['REFERENCED_COLUMN_NAME']
            parents_cst = cst[
                (cst['TABLE_NAME'] != self.root) &
                (cst['REFERENCED_TABLE_NAME'] != parent_root)
            ]
            parent = Tree(
                con=con,
                tables=self.tables,
                root=parents_cst_row['REFERENCED_TABLE_NAME'],
                cst=parents_cst,
                ref=parent_ref,
                reffed=parent_reffed
            )
            self.parents.append(parent)

        self.parents.sort(key=lambda x: self.tables[x.root].order)

        for i, child_cst_row in self.reffed_info.iterrows():
            child_root = child_cst_row['TABLE_NAME']
            child_ref = child_cst_row['COLUMN_NAME']
            child_reffed = child_cst_row['REFERENCED_COLUMN_NAME']
            children_cst = cst[
                (cst['REFERENCED_TABLE_NAME'] != self.root) &
                (cst['TABLE_NAME'] != self.root)
                ]
            child = Tree(
                con=con,
                tables=self.tables,
                root=child_root,
                cst=children_cst,
                ref=child_ref,
                reffed=child_reffed
            )
            self.children.append(child)

        self.children.sort(key=lambda x: self.tables[x.root].order)
        self.booking_sequence = self.get_booking_sequence()
        self.reading_sequence = self.booking_sequence.__reversed__()

        self.node_names = self.get_node_names()
        self.sort_cols()

    def __bool__(self):
        if self.root:
            return True
        else:
            return False

    def __gt__(self, other):
        return self.root in [p for p in other.all_parenthood_names()]

    def __lt__(self, other):
        return self.root in [c for c in other.all_childhood_names()]

    def __ge__(self, other):
        return self.root in [p for p in other.all_parenthood_names()] or self.root == other.root

    def __le__(self, other):
        return self.root in [c for c in other.all_childhood_names()] or self.root == other.root

    def __eq__(self, other):
        return self.root == other.root

    def get_booking_sequence(self):
        res = sum([
            p.get_booking_sequence() for p in self.parents
        ], start=[]) + [
            self.root
        ] + sum([
            c.get_booking_sequence() for c in self.children
        ], start=[])

        return [item for i, item in enumerate(res) if item not in res[:i]]

    def all_parenthood_names(self):
        res = set([p.root for p in self.parents])
        for p in self.parents:
            res |= p.all_parenthood_names()
        for c in self.children:
            res |= c.all_parenthood_names()
        return res

    def all_childhood_names(self):
        res = set([c.root for c in self.children])
        for c in self.children:
            res |= c.all_parenthood_names()
        return res

    def get_child_path(self, child_root: str):
        res = []
        for c in self.children:
            if c.root == child_root:
                return [self.root, c.root]

            res.extend(c.get_child_path(child_root=child_root))
        if len(res) > 0:
            res = [self.root] + res

        return res

    def sort_cols(self):
        cols = self.table.cols
        col_orders = dict(zip(cols.keys(), [col.order for col in cols.values()]))

        p_orders = {parent.ref: cols[parent.ref].order for parent in self.parents}
        sorted_orders = sorted([i for i in p_orders.values()])
        p_orders = {key: sorted_orders.index(value) for key, value in p_orders.items()}

        col_orders = {key: value for key, value in col_orders.items()
                      if key not in p_orders.keys()}

        sorted_orders = sorted([i for i in col_orders.values()])
        col_orders = {key: sorted_orders.index(value) + len(p_orders) for key, value in col_orders.items()}

        for key, order in p_orders.items():
            self.table.cols[key].order = order

        for key, order in col_orders.items():
            self.table.cols[key].order = order

    def json_obj_base(self):
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'pk', 'ref', 'reffed', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        for parent in self.parents:
            json_obj['parents'].append(parent.json_obj_base())
        for child in self.children:
            json_obj['children'].append(child.json_obj_base())
        json_obj['table'] = self.table.to_json_obj()
        return json_obj

    @property
    def json_obj(self):
        return self.json_obj_base()

    def p_(self, p_name, ref=None):
        if not ref:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name
                ), None
            )
        else:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name and parent.ref == ref
                ), None
            )

    def c_(self, c_name, reffed=None):
        if not reffed:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name
                ), None
            )
        else:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name and child.reffed == reffed
                ), None
            )

    @property
    def p(self):
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        return PropertyIndexer(self, 'c_')

    def get_all_nodes(self, res=None):
        if not res:
            res = {self.root: self}
        for p in self.parents:
            if p.root not in list(res.keys()):
                res[p.root] = p
                res.update(p.get_all_nodes(res=res))
        for c in self.children:
            if c.root not in list(res.keys()):
                res[c.root] = c
            res.update(c.get_all_nodes(res=res))
        return res

    def get_all_parents(self, res=None):
        if not res:
            res = {}
        for p in self.parents:
            if p.root not in list(res.keys()):
                res[p.root] = p
                res.update(p.get_all_parents(res=res))
        for c in self.children:
            res.update(c.get_all_parents(res=res))
        return res

    def get_node_names(self):
        res = {self.root}
        for parent in self.parents:
            res |= parent.node_names
        for child in self.children:
            res |= child.node_names
        return res

    def get_parent_tree_structure(self):
        res = {p.root: p.get_parent_tree_structure() for p in self.parents}
        return res


class DataTree(Tree):
    def __init__(
            self,
            con=None,
            tables=None,
            tree=None,
            root=None,
            cst_pki=None,
            ref=None,
            reffed=None,
            relevant_data_set=None,
    ):
        if con is None:
            if tree is None:
                print('con and tree cannot be None at the same time')
                raise ValueError
            else:
                con = copy(tree.con)
        if tables is None:
            if tree is None:
                print('tables and tree cannot be None at the same time')
                raise ValueError
            else:
                tables = copy(tree.tables)

        self.schemas = [table.schema for table in tables.values()]

        if cst_pki is None:
            cst_pki = get_cst_pki(con=con, schemas=self.schemas)

        super().__init__(
            con=con,
            tree=tree,
            root=root,
            cst_pki=cst_pki,
            tables=tables,
            ref=ref,
            reffed=reffed,
        )

        if relevant_data_set is None:
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        try:
            self.data = relevant_data_set[self.root]
        except KeyError:
            self.data = pd.DataFrame(columns=self.table.cols.keys())
        self.relevant_data_set = relevant_data_set

    def __len__(self):
        return len(self.data)

    def __getitem__(self, key):
        return self.data[key]

    @property
    def loc(self):
        return self.data.loc

    @property
    def iloc(self):
        return self.data.iloc

    def iterrows(self):
        return self.data.iterrows()

    def itertrees(self):
        for i, row in self.iterrows():
            relevant_data_set = deepcopy(self.relevant_data_set)
            relevant_data_set[self.root] = row.to_frame().transpose()
            datatree = DataTree(
                con=self.con,
                tables=self.tables,
                tree=self,
                relevant_data_set={}
            )
            datatree.from_relevant_data_set(relevant_data_set)
            yield i, datatree

    @staticmethod
    def update_values_map(values_map, tree, root_data):
        for p in tree.parents:
            if len(root_data) == 0:
                adding_data = set()
            else:
                try:
                    adding_data = set(root_data[p.ref].dropna())
                except KeyError:
                    print(f'parent ref info error: {p.ref}')
                    adding_data = set()

            try:
                values_map[p.root][p.reffed]['values'] = \
                    values_map[p.root][p.reffed]['values'] | adding_data
            except KeyError:
                values_map[p.root] = {}
                values_map[p.root][p.reffed] = {
                    'tree': p,
                    'values': adding_data
                }

        for c in tree.children:
            if len(root_data) == 0:
                adding_data = set()
            else:
                adding_data = set(root_data[c.reffed].dropna())

            try:
                values_map[c.root][c.ref]['values'] = \
                    values_map[c.root][c.ref]['values'] | adding_data
            except KeyError:
                values_map[c.root] = {}
                values_map[c.root][c.ref] = {
                    'tree': c,
                    'values': adding_data
                }

        return values_map

    def fill_na_with_none(self):
        relevant_data_set_copy = deepcopy(self.relevant_data_set)
        for name, df in relevant_data_set_copy.items():
            df = df.fillna(np.nan).replace([np.nan], [None])
            self.relevant_data_set[name] = df

    def from_sql(
            self,
            index_col: str = None,
            index_values: Set[str] = (),
            limit=None,
            offset=None
    ):
        # root data
        root_schema = self.table.schema
        if index_col is None:
            where_str = ''
        else:
            values_str = ', '.join([f'"{index_value}"' for index_value in index_values])
            where_str = f'WHERE `{index_col}` in ({values_str})'

        if limit is None or offset is None:
            limit_offset_str = ''
        else:
            limit_offset_str = f'limit {str(limit)} offset {str(offset)}'

        sql = f'SELECT * FROM {root_schema}.{self.root} {where_str} {limit_offset_str}'
        root_data = pd.read_sql(sql=sql, con=self.con, index_col=self.pk)
        if len(root_data) == 0:
            print(f'empty result for root: "{self.root}" \n'
                  f'with index col: "{where_str}" \n'
                  f'with values in {index_values}')
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        else:
            relevant_data_set = {self.root: root_data}

            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # relevant data
            reading_seq = list(self.reading_sequence)
            reading_seq.remove(self.root)

            for node_name in reading_seq:
                where_str_list = []
                if node_name not in values_map:
                    continue
                for col, tree_values in values_map[node_name].items():
                    values = tree_values['values']

                    # has no value by col
                    if len(values) == 0:
                        continue

                    values_str = ', '.join(['"%s"' % value.replace('%', '%%') for value in values])
                    where_str = f'(`{col}` in ({values_str}))'
                    where_str_list.append(where_str)

                # has no value in table
                if len(where_str_list) == 0:
                    continue

                all_where_str = ' or '.join(where_str_list)

                sql = f'SELECT * FROM {self.tables[node_name].schema}.{node_name} WHERE {all_where_str}'
                data = pd.read_sql(sql=sql, con=self.con, index_col='id')
                relevant_data_set[node_name] = data

                for col, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

        self.data = relevant_data_set[self.root]
        self.relevant_data_set = relevant_data_set

    def from_relevant_data_set(self, relevant_data_set=None):
        relevant_data_set_copy = copy(relevant_data_set)

        if relevant_data_set_copy is None:
            relevant_data_set_copy = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))

        root_data = relevant_data_set_copy[self.root]
        if len(root_data) == 0:
            relevant_data_set_copy = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        else:
            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # relevant data
            reading_seq = self.reading_sequence[1:]

            for node_name in reading_seq:
                if node_name not in relevant_data_set_copy:
                    relevant_data_set_copy[node_name] = pd.DataFrame(
                        columns=[col_name for col_name, _ in self.tables[node_name].cols.items()]
                    )
                    continue
                if node_name not in values_map:
                    continue
                where_str_list = []
                data = pd.DataFrame(columns=self.table.cols.keys())
                for col, tree_values in values_map[node_name].items():
                    values = tree_values['values']
                    values_str = ', '.join(['"%s"' % value for value in values])
                    data = relevant_data_set_copy[node_name]
                    where_str = f'(data["{col}"].isin([{values_str}]))'
                    where_str_list.append(where_str)

                all_where_str = ' | '.join(where_str_list)

                relevant_data_set_copy[node_name] = eval(f'data[{all_where_str}]')

                for col, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

        self.data = root_data
        self.relevant_data_set = relevant_data_set_copy
        self.fill_na_with_none()

    @check_param_valid_range(
        [
            'if_data_exists'
        ],
        [
            ['fail', 'skip', 'merge', 'update']
        ]
    )
    def to_sql(self, curd, if_data_exists='fail'):
        stmt = curd.stmt_df_merge(
            datatree=self,
            if_data_exists=if_data_exists,
        )
        print(stmt)
        # con.cursor.execute(stmt)

    def p_(self, p_name, ref=None):
        p_tree = Tree.p_(self, p_name=p_name, ref=ref)
        p = DataTree(
            con=self.con,
            tables=self.tables,
            tree=p_tree
        )
        p.from_relevant_data_set(self.relevant_data_set)
        return p

    def c_(self, c_name, reffed=None):
        c = DataTree(
            con=self.con,
            tables=self.tables,
            tree=Tree.c_(self, c_name=c_name, reffed=reffed)
        )
        c.from_relevant_data_set(self.relevant_data_set)
        return c

    @property
    def p(self):
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        return PropertyIndexer(self, 'c_')

    @property
    def ps(self):
        for parent in self.parents:
            p = DataTree(
                con=self.con,
                tables=self.tables,
                tree=parent
            )
            p.from_relevant_data_set(self.relevant_data_set)
            yield p

    @property
    def cs(self):
        for child in self.children:
            c = DataTree(
                con=self.con,
                tables=self.tables,
                tree=child
            )
            c.from_relevant_data_set(self.relevant_data_set)
            yield c

    def json_obj_base(self, with_value=False):
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'pk', 'ref', 'reffed', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        data = self.data.copy(deep=True)
        data.replace(np.nan, None, inplace=True)
        for parent in self.ps:
            json_obj['parents'].append(parent.json_obj_base(with_value=with_value))
        for child in self.cs:
            json_obj['children'].append(child.json_obj_base(with_value=with_value))

        if with_value:
            json_obj['values'] = data.to_dict(orient='list')

        json_obj['table'] = self.table.to_json_obj()

        return json_obj

    @property
    def json_obj(self):
        return self.json_obj_base(with_value=True)

    # @profile_line_by_line
    def nested_values(self, ref_group=None, ignore_ref_col=None, full_detail=False):
        if full_detail:
            col_items = [
                item for item in self.table.cols.items()
                if not item[1].col_name == ignore_ref_col
            ]
        else:
            col_items = [
                item for item in self.table.cols.items()
                if not pd.isna(item[1].web_list_order)
                and not item[1].col_name == ignore_ref_col
            ]
        col_items.sort(key=lambda x: x[1].web_list_order)
        display_column_names, display_columns = zip(*col_items)
        display_column_names = ['key'] + list(display_column_names)
        columns = [
            {
                'key': i,
                'dataIndex': col.col_name,
                'title': col.label,
                'dataType': col.data_type
            } for i, col in enumerate(display_columns)
        ]
        data = self.data.assign(key=self.data.index.astype(str))
        data['id'] = data.index
        data.sort_index()
        display_data = data[display_column_names]

        if not ref_group:
            data_source = {0: display_data.to_dict(orient='records')}
        else:
            data_source = {
                key: display_data.loc[names].to_dict(orient='records')
                for key, names in ref_group.items()
            }

        children = {}
        for child in self.cs:
            child_data = child.data.sort_index()
            gs_data = {key: value for key, value in data.groupby(by=child.reffed)}
            ref_group = {
                str(gs_data[ref_value].index[0]): group.index.to_list()
                for ref_value, group in child_data.groupby(by=child.ref)
            }
            children[child.root] = child.nested_values(
                ref_group=ref_group,
                ignore_ref_col=child.ref,
                full_detail=full_detail
            )

        res = {
            'tag': self.table.label,
            'columns': columns,
            'dataSource': data_source,
            'children': children
        }

        return res

    # @profile_line_by_line
    def nested_values2(self, data=None, parent_ref=None):
        if not isinstance(data, pd.DataFrame):
            data = self.data
        col_items = [
            item for item in self.table.cols.items()
            if not pd.isna(item[1].web_list_order)
            and item[1].col_name != parent_ref
        ]
        col_items.sort(key=lambda x: x[1].web_list_order)
        display_column_names, display_columns = zip(*col_items)
        display_column_names = ['key'] + list(display_column_names)
        columns = [
            {
                'key': i,
                'dataIndex': col.col_name,
                'title': col.label,
                'dataType': col.data_type
            } for i, col in enumerate(display_columns)
        ]
        data = data.assign(key=data.index.astype(str))
        data['id'] = data.index
        data.sort_index()

        if not parent_ref:
            display_data = data[display_column_names]
            data_source = {0: display_data.to_dict(orient='records')}
        else:
            gs_data = data.groupby(by=parent_ref)
            data_source = {
                key: g_data[display_column_names].to_dict(orient='records')
                for key, g_data in gs_data
            }

        children = {}
        for child in self.cs:
            parent_data = pd.DataFrame(
                data=data[['id', child.reffed]]
            )
            parent_data.columns = [f'id_from_{self.root}', child.reffed]
            child_data = pd.merge(left=child.data,
                                  right=parent_data,
                                  left_on=child.ref, right_on=child.reffed,
                                  suffixes=('', f'_from_{self.root}'))
            children[child.root] = child.nested_values2(
                data=child_data,
                parent_ref=f'id_from_{self.root}'
            )

        res = {
            'tag': self.table.label,
            'columns': columns,
            'dataSource': data_source,
            'children': children
        }

        return res

    def get_all_parents_with_full_value(self):
        res = {}
        all_parents = self.get_all_parents()
        for p_name, p in all_parents.items():
            dp = DataTree(
                con=self.con,
                tables=self.tables,
                tree=p
            )
            dp.from_sql()
            res[p.root] = dp
        return res

    def get_parents_select_values(self):
        res = {}
        all_parents = self.get_all_parents_with_full_value()
        for p_name, p in all_parents.items():
            try:
                res[p.root] = p.data[p.reffed].to_list()
            except KeyError:
                if p.data.index.name == p.reffed:
                    res[p.root] = p.data.index.to_list()
                else:
                    raise KeyError

        return res

    def from_excel_booking_sheet(self, dfs):
        relevant_data_set = {}
        root_sheet_name = f'bks_{self.table.label}'
        root_df_raw = dfs[root_sheet_name]
        root_df = get_crop_from_df(
            df=root_df_raw,
            anchor_x=3,
            anchor_y=4,
            vertical=True,
            col_offset=-2,
            pk_index=None
        )

        relevant_data_set[self.root] = root_df

        for child in self.children:
            anchor_x = root_df_raw.iloc[:, 2].to_list().index(child.root) + 4
            child_df = get_crop_from_df(
                df=root_df_raw,
                anchor_x=anchor_x,
                anchor_y=3,
                vertical=False,
                col_offset=-2,
                pk_index=-3
            )
            # mark child ref
            child_df[child.ref] = root_df[child.reffed].values[0]

            # mark child auto-name
            naming_cols = sorted([
                col_obj for col_name, col_obj in child.table.cols.items()
                if not pd.isna(col_obj.naming_field_order)
            ], key=lambda x: x.naming_field_order)
            if len(naming_cols) > 0:
                child_df['name'] = child_df.apply(
                    lambda x: '-'.join([
                        x[col_obj.col_name] for col_obj in naming_cols
                    ]),
                    axis=1
                )

            relevant_data_set[child.root] = child_df

        for parent_root, parent in self.get_all_parents().items():
            try:
                parent_sheet_name = f'bks_{parent.table.label}'
            except KeyError:
                print(f'no sheet for root "{parent_root}"')
                continue

            try:
                parent_df_raw = dfs[parent_sheet_name]
            except KeyError:
                continue
            parent_df = get_crop_from_df(
                df=parent_df_raw,
                anchor_x=7,
                anchor_y=3,
                vertical=False,
                col_offset=-2,
                pk_index=-3
            )

            # mark auto name
            naming_cols = [(col_name, col.naming_field_order)
                             for col_name, col in parent.table.cols.items()
                             if not pd.isna(col.naming_field_order)]
            if len(naming_cols) > 0:
                naming_cols = [col for col, order in sorted(naming_cols, key=lambda x: x[1])]

                if len(parent_df) > 0:
                    if 'name' in parent_df.columns:
                        parent_df['name'] = parent_df.apply(
                            lambda x: '-'.join([x[col] for col in naming_cols]) if pd.isna(x['name']) else x[
                                'name'],
                            axis=1
                        )
                    else:
                        parent_df['name'] = parent_df.apply(
                            lambda x: '-'.join([x[col] for col in naming_cols]),
                            axis=1
                        )

            relevant_data_set[parent_root] = parent_df

        for key, df in relevant_data_set.items():
            relevant_data_set[key] = df.where(df.notnull(), None)

        self.from_relevant_data_set(relevant_data_set=relevant_data_set)

    def to_excel_sheets(self, pth, node_names=None, web_visible_only=False, label=False, table_label=False):
        writer = ExcelWriter(pth, mode='w', engine='openpyxl', options={'strings_to_urls': False})
        if not node_names:
            node_names = self.node_names
        flag = False

        for node_name in node_names:
            table = self.tables[node_name]
            if web_visible_only:
                columns = [
                    col.col_name
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            else:
                columns = [
                    col.col_name
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            try:
                df = self.relevant_data_set[node_name][columns]
            except KeyError:
                df = pd.DataFrame(columns=columns)

            if label:
                column_mapper = dict(
                    [
                        (col.col_name, col.label)
                        for col in table.cols.values()]
                )
                df.rename(columns=column_mapper, inplace=True)
            if table_label:
                node_name = table.label

            df.to_excel(writer, sheet_name=node_name, index=False,
                        columns=df.columns)
            fit_col_width(writer=writer, df=df, sheet_name=node_name)
            flag = True

        if flag:
            writer._save()
            writer.close()
        else:
            del writer
            os.remove(pth)

    def to_excel_detail_sheet(self, pth):
        child_data = {}
        for child in self.children:
            child_data[child.root] = self[child.root].to_dict()
        wb = Workbook()
        ws = wb.worksheets[0]
        pin = 1
        for r, (key, value) in enumerate(self.data.items()):
            if self.table.cols[key].web_visible == 1:
                ws.cell(row=pin, column=1).value = self.table.cols[key].label
                ws.cell(row=pin, column=2).value = value
                pin += 1

        pin += 1

        for child_name, child_values in child_data.items():
            ws.cell(row=pin, column=1).value = self.tables[child_name].comment
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    ws.cell(row=pin, column=c_index).value = self.tables[child_name].cols[col].label
                    c_index += 1
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    for r, (index, v) in enumerate(value.items()):
                        ws.cell(row=pin + r, column=c_index).value = v
                    c_index += 1
            pin += len(child_values) + 1

        wb.save(pth)


if __name__ == '__main__':
    pass
