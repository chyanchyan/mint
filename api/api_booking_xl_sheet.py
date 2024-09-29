import shutil
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell import Cell
import pandas as pd
import os
import sys

from mint.db.utils import get_con
from mint.sys_init import TABLES

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)

if parent_dir not in sys.path:
    sys.path.append(parent_dir)

from mint.db.tree import DataTree, Tree


def apply_cell_format(cell_src: Cell, cell_target: Cell):
    cell_src._style = cell_target._style


def apply_conditional_format(ws_src, cell_src, ws_target, cell_target):
    cf_rules: dict = ws_target.conditional_formatting._cf_rules
    # 遍历规则字典
    for key, item in cf_rules.items():
        t_coord = cell_target.coordinate
        if t_coord == list(key.cells.ranges)[0].coord:
            rule = cf_rules[key][0]
            ws_src.conditional_formatting.add(cell_src.coordinate, rule)
            break


def apply_data_validation(sheet, cell, col, formula1):

    dv = DataValidation(
        type="list",
        formula1=formula1,
        showDropDown=False,
        allowBlank=False,
        error=f'该"{col.label}"不存在，请重新输入或选择，或前往网页新增选项',
        errorTitle='输入值不存在',
        prompt='请选择列表内值',
        promptTitle='请选择列表内值',
        showErrorMessage=True
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def apply_radio_validation(sheet, cell):
    dv = DataValidation(
        type="list",
        formula1='"1, 0"',
        showDropDown=False,
        allowBlank=False,
        error=f'请输入: 0 - 否，1 - 是',
        errorTitle='输入值不合法',
        prompt='请选择：0 - 否，1 - 是',
        promptTitle='请选择：0 - 否，1 - 是',
        showErrorMessage=True
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def apply_number_validation(sheet, cell):
    dv = DataValidation(
        type="decimal",
        allowBlank=False,
        error=f'请输入数字',
        errorTitle='输入值不合法',
        prompt='请输入数字',
        promptTitle='请输入数字',
        showErrorMessage=True,
        operator='between'
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def fill_table(
        root,
        ws_booking,
        ws_cell_format,
        cell_formats,
        select_values,
        dst_row,
        dst_col,
        table,
        tables,
        direction='vertical',
        values=None,
        is_selected_values=False,
        start_value_row_index=0,
        show_non_display_name=False,
        fill_headers=True,
        value_row_offset=0
):

    parent_table_data_start_row = 8

    if not show_non_display_name:
        col_list = [
            col for col in table.cols
            if col.web_visible or
            col.table_name == 'auto_name'
        ]
    else:
        col_list = [
            col for col in table.cols
            if col.web_visible
        ]

    col_list = [
        col for col in col_list
        if (not col.foreign_key or  # 引用列为None，或
           pd.isna(col.foreign_key) or # 引用列为nan，或
           col.foreign_key.split('.')[1] != root)
        or not (col.naming_field_order is None or pd.isna(col.naming_field_order)) # 排除命名字段
    ]

    # fill table label
    if fill_headers:
        apply_cell_format(
            cell_src=ws_booking.cell(row=dst_row, column=dst_col + 1, value=table.label),
            cell_target=cell_formats['table_label']
        )

        # fill table name
        if direction == 'vertical':
            ws_booking.cell(row=dst_row - 1, column=dst_col + 2, value=table.table_name)
        elif direction == 'horizontal':
            ws_booking.cell(row=dst_row, column=dst_col + 2, value=table.table_name)

    # fill root cols
    value_row_index = int(start_value_row_index)
    for col_idx, col in enumerate(col_list):
        is_nullable = (
            col.nullable == 1 and
            (
                col.foreign_key is None or
                pd.isna(col.foreign_key)
            )
        )
        col_label = ['*', ''][is_nullable] + col.label
        # cell validation
        if col.foreign_key and not pd.isna(col.foreign_key):
            db_name, parent_table_name, fk = col.foreign_key.split('.')
            if parent_table_name == root:   # 若为子表引用根表，则引用根表数据
                formula1 = (
                    f'=bks_{tables[parent_table_name].label}!$E${[
                        col.col_name 
                        for col in filter(lambda x: x.web_visible == 1, tables[parent_table_name].cols)
                    ].index(fk) + 4}'
                )
            else:
                parent_table_obj = tables[parent_table_name]
                select_values_rows_count = len(select_values[parent_table_name])
                parent_visible_col_names = [
                    col.col_name for col in tables[parent_table_name].cols
                    if col.web_visible or col.table_name == 'auto_name'
                ]
                try:
                    fk_idx = parent_visible_col_names.index(fk) + 4
                except ValueError:
                    print(fk)
                    raise ValueError
                col_letter = get_column_letter(fk_idx)

                formula1 = f'=bks_{parent_table_obj.label}!${col_letter}${parent_table_data_start_row}:' \
                           f'${col_letter}${str(select_values_rows_count + parent_table_data_start_row + 1)}'

        if direction == 'vertical':
            cell_col = ws_booking.cell(row=dst_row + col_idx, column=dst_col + 3, value=col_label)
            ws_booking.cell(row=dst_row + col_idx, column=dst_col + 2, value=col.col_name)
            cell_default = ws_booking.cell(row=dst_row + col_idx, column=dst_col + 4)

            if col.table_name == 'auto_name':
                auto_naming_cols = sorted([
                    (i, col)
                    for i, col in enumerate(col_list)
                    if not (pd.isna(col.naming_field_order) or col.naming_field_order is None)
                ], key=lambda x: x[1].naming_field_order)
                auto_naming_col_addrs = [f'E{dst_row + item[0]}' for item in auto_naming_cols]
                default_value = '=' + '&"-"&'.join(auto_naming_col_addrs)
                apply_cell_format(cell_default, cell_formats['auto_name'])
            else:
                default_value = col.default
            cell_default.value = default_value

            if col.foreign_key and not pd.isna(col.foreign_key):
                db_name, foreign_table, foreign_key = col.foreign_key.split('.')
                ws_booking.cell(
                    row=dst_row + col_idx,
                    column=dst_col + 5,
                    value='fk=' + tables[foreign_table].label
                )
            if values:
                if len(values[col.col_name]) > 0:
                    cell_default.value = values[col.col_name][0]

        elif direction == 'horizontal':
            cell_default = ws_booking.cell(row=dst_row + 2, column=dst_col + col_idx + 3)

            if fill_headers:
                cell_col = ws_booking.cell(row=dst_row, column=dst_col + col_idx + 3, value=col_label)
                ws_booking.cell(row=dst_row + 1, column=dst_col + col_idx + 3, value=col.col_name)

                if col.table_name == 'auto_name':
                    auto_naming_cols = sorted([
                        (i, col)
                        for i, col in enumerate(col_list)
                        if not (pd.isna(col.naming_field_order) or col.naming_field_order is None)
                    ], key=lambda x: x[1].naming_field_order)
                    auto_naming_col_addrs = [f'{get_column_letter(dst_col + item[0] + 3)}{dst_row + 2}' for item in auto_naming_cols]
                    default_value = '=' + '&"-"&'.join(auto_naming_col_addrs)
                    apply_cell_format(cell_default, cell_formats['auto_name'])
                else:
                    if not pd.isna(col.foreign_key) and col.foreign_key is not None:
                        db_name, foreign_table, foreign_key = col.foreign_key.split('.')
                        if foreign_table != root:
                            default_value = col.default
                        else:
                            default_value = formula1
                    else:
                        default_value = col.default
                cell_default.value = default_value

                if col.foreign_key and not pd.isna(col.foreign_key):
                    db_name, foreign_table, foreign_key = col.foreign_key.split('.')
                    if foreign_table != root:
                        ws_booking.cell(
                            row=dst_row - 1,
                            column=dst_col + col_idx + 3,
                            value='fk=' + tables[foreign_table].label
                        )
            else:
                cell_col = ws_booking.cell(row=dst_row, column=dst_col + col_idx + 3)

            if values:
                try:
                    col_values = values[col.col_name]

                    for r, value in enumerate(col_values.values()):
                        if fill_headers:
                            row = dst_row + value_row_offset + 2 + r + 1
                        else:
                            row = dst_row + value_row_offset + r + 1

                        cell_value = ws_booking.cell(row=row, column=dst_col + col_idx + 3, value=value)
                        if not is_selected_values:
                            ws_booking.cell(row=row, column=dst_col, value=f'row{value_row_index}')
                            apply_cell_format(
                                cell_src=cell_value,
                                cell_target=cell_formats[col.web_obj]
                            )
                            apply_conditional_format(
                                ws_src=ws_booking,
                                cell_src=cell_value,
                                ws_target=ws_cell_format,
                                cell_target=cell_formats[col.web_obj]
                            )

                            if col.foreign_key and not pd.isna(col.foreign_key):
                                db_name, foreign_table, foreign_key = col.foreign_key.split('.')
                                if foreign_table != root:
                                    apply_data_validation(sheet=ws_booking, cell=cell_value, col=col,
                                                          formula1=formula1)
                                else:
                                    ws_booking.cell(row=row, column=dst_col + col_idx + 3, value=formula1)
                            if col.web_obj == 'radio':
                                apply_radio_validation(
                                    sheet=ws_booking, cell=cell_value
                                )
                            if col.table_name == 'auto_name':
                                apply_cell_format(cell_value, cell_formats['auto_name'])

                            value_row_index += 1
                except KeyError:
                    pass

        else:
            raise ValueError

        if col.table_name == 'auto_name':
            apply_cell_format(
                cell_src=cell_col,
                cell_target=cell_formats['auto_name'])
            apply_cell_format(
                cell_src=cell_default,
                cell_target=cell_formats['auto_name']
            )
        else:
            apply_cell_format(
                cell_src=cell_col,
                cell_target=cell_formats['column'])
            apply_cell_format(
                cell_src=cell_default,
                cell_target=cell_formats[col.web_obj]
            )

        # cell conditional format
        apply_conditional_format(
            ws_src=ws_booking,
            cell_src=cell_default,
            ws_target=ws_cell_format,
            cell_target=cell_formats[col.web_obj]
        )

        # cell validation
        if col.foreign_key and not pd.isna(col.foreign_key):
            apply_data_validation(
                sheet=ws_booking,
                cell=cell_default,
                col=col,
                formula1=formula1
            )

        if col.web_obj == 'radio':
            apply_radio_validation(
                sheet=ws_booking,
                cell=cell_default
            )

    return value_row_index


def render_booking_xl_sheet(output_path, template_path, data_tree: DataTree):

    tables = data_tree.tables
    root = data_tree.root
    select_values = data_tree.get_parents_select_values()

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path, keep_vba=True)
    ws_cell_format = wb['cell_format']
    ws_booking = wb.copy_worksheet(wb['bks_root'])
    ws_booking.title = 'bks_' + data_tree.table.label

    data_cell_format = pd.read_excel(template_path, sheet_name='cell_format')
    cell_formats = defaultdict(lambda: ws_cell_format.cell(row=4, column=2))
    d = dict(zip(
        data_cell_format['cell_type'].values,
        [ws_cell_format.cell(row=r + 2, column=2)
         for r in range(len(data_cell_format['cell_type']))]
    ))
    for k, v in d.items():
        cell_formats[k] = v

    dst_row = 4
    dst_col = 1
    value_row_index = 0

    # fill parent tables
    parent_trees = data_tree.get_all_parents_with_values()
    parents_trees_full_value = data_tree.get_all_parents_with_full_value()

    for parent_name, parent in parent_trees.items():
        ws_parent_booking = wb.copy_worksheet(wb['bks_root'])
        ws_parent_booking.title = 'bks_' + parent.table.label
        parent_full_value = parents_trees_full_value[parent.root]
        value_row_index = fill_table(
            root=root,
            ws_booking=ws_parent_booking,
            ws_cell_format=ws_cell_format,
            cell_formats=cell_formats,
            select_values=select_values,
            dst_row=5,
            dst_col=1,
            table=parent.table,
            tables=tables,
            direction='horizontal',
            start_value_row_index=value_row_index,
            values=parent.data.reset_index().to_dict(),
            show_non_display_name=False
        )

        fill_table(
            root=root,
            ws_booking=ws_parent_booking,
            ws_cell_format=ws_cell_format,
            cell_formats=cell_formats,
            select_values=select_values,
            dst_row=5,
            dst_col=1,
            table=parent.table,
            tables=tables,
            direction='horizontal',
            values={parent.reffed: parent_full_value.data.reset_index().to_dict()[parent.reffed]},
            is_selected_values=True,
            show_non_display_name=False,
            fill_headers=False,
            value_row_offset=len(parent.data)
        )

        ws_parent_booking.row_dimensions[6].hidden = True
        ws_parent_booking.row_dimensions[7].hidden = True
        ws_parent_booking.column_dimensions['C'].hidden = True

    # fill root table
    # fill table label
    table = tables[root]
    fill_table(
        root=root,
        ws_booking=ws_booking,
        ws_cell_format=ws_cell_format,
        cell_formats=cell_formats,
        select_values=select_values,
        dst_row=dst_row,
        dst_col=dst_col,
        table=table,
        tables=tables,
        direction='vertical',
        values=data_tree.data.to_dict(orient='list'),
        show_non_display_name=False
    )

    dst_row = ws_booking.max_row + 3

    # fill child tables
    for child in data_tree.cs:
        value_row_index = fill_table(
            root=root,
            ws_booking=ws_booking,
            ws_cell_format=ws_cell_format,
            cell_formats=cell_formats,
            select_values=select_values,
            dst_row=dst_row,
            dst_col=dst_col,
            table=child.table,
            direction='horizontal',
            tables=tables,
            values=child.data.to_dict(),
            start_value_row_index=value_row_index,
            show_non_display_name=True
        )

        # hide format row
        ws_booking.row_dimensions[dst_row + 1].hidden = True
        ws_booking.row_dimensions[dst_row + 2].hidden = True

        dst_row = ws_booking.max_row + 3

    ws_booking.column_dimensions['C'].hidden = True

    wb.save(output_path)


if __name__ == '__main__':
    con = get_con('data')
    t = Tree(
        con=con,
        root='project',
        tables=TABLES,
    )
    for item in list(t.get_all_parents().keys()):
        print(item)