import shutil
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell import Cell

if 'mint' in __name__.split('.'):
    from .sys_init import *
    from .tree import DataTree
else:
    from sys_init import *
    from tree import DataTree


def apply_cell_format(cell_src: Cell, cell_target: Cell):
    cell_src._style = cell_target._style


def apply_conditional_format(ws_src, cell_src, ws_target, cell_target):
    cf_rules: dict = ws_target.conditional_formatting._cf_rules
    # 遍历规则字典
    for key, item in cf_rules.items():
        t_coord = cell_target.coordinate
        if t_coord == list(key.cells.ranges)[0].coord:
            rule = cf_rules[key][0]
            ws_src.conditional_formatting.add(cell_src.coordinate, rule)
            break


def apply_data_validation(sheet, cell, col_obj, formula1):

    dv = DataValidation(
        type="list",
        formula1=formula1,
        showDropDown=False,
        allowBlank=False,
        error=f'该"{col_obj.label}"不存在，请重新输入或选择，或前往网页新增选项',
        errorTitle='输入值不存在',
        prompt='请选择列表内值',
        promptTitle='请选择列表内值',
        showErrorMessage=True
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def apply_radio_validation(sheet, cell):
    dv = DataValidation(
        type="list",
        formula1='"1, 0"',
        showDropDown=False,
        allowBlank=False,
        error=f'请输入: 0 - 否，1 - 是',
        errorTitle='输入值不合法',
        prompt='请选择：0 - 否，1 - 是',
        promptTitle='请选择：0 - 否，1 - 是',
        showErrorMessage=True
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def apply_number_validation(sheet, cell):
    dv = DataValidation(
        type="decimal",
        allowBlank=False,
        error=f'请输入数字',
        errorTitle='输入值不合法',
        prompt='请输入数字',
        promptTitle='请输入数字',
        showErrorMessage=True,
        operator='between'
    )

    dv.add(cell)
    sheet.add_data_validation(dv)


def fill_table(
        root,
        ws_booking,
        ws_cell_format,
        cell_formats,
        select_values,
        dst_row,
        dst_col,
        table,
        tables,
        direction='vertical',
        values=None,
        is_selected_values=False,
        start_value_row_index=0,
        show_display_name=True
):

    # get col list
    if show_display_name:
        col_list = sorted([
            col_name for col_name in table.cols
            if table.cols[col_name].web_visible or
            table.cols[col_name].table_name == 'auto_name'
        ], key=lambda x: table.cols[x].order)
    else:
        col_list = sorted([
            col_name for col_name in table.cols
            if table.cols[col_name].web_visible
        ], key=lambda x: table.cols[x].order)

    col_list = [
        col_name for col_name in col_list
        if not table.cols[col_name].foreign_key or
           pd.isna(table.cols[col_name].foreign_key) or
           table.cols[col_name].foreign_key.split('.')[1] != root
    ]

    # fill table label
    apply_cell_format(
        cell_src=ws_booking.cell(row=dst_row, column=dst_col + 1, value=table.label),
        cell_target=cell_formats['table_label']
    )

    # fill table name
    if direction == 'vertical':
        ws_booking.cell(row=dst_row - 1, column=dst_col + 2, value=table.table_name)
    elif direction == 'horizontal':
        ws_booking.cell(row=dst_row, column=dst_col + 2, value=table.table_name)

    # fill root cols
    value_row_index = int(start_value_row_index)
    for col_idx, col_name in enumerate(col_list):
        col_obj = table.cols[col_name]
        col_label = ['*', ''][col_obj.nullable == 1 and col_obj.foreign_key is None] + col_obj.label
        # cell validation
        if col_obj.foreign_key and not pd.isna(col_obj.foreign_key):
            db_name, parent_table_name, fk = col_obj.foreign_key.split('.')
            if parent_table_name == root:
                continue
            parent_table_obj = tables[parent_table_name]
            select_values_rows_count = len(select_values[parent_table_name])
            parent_visible_col_names = [
                col.col_name for col in tables[parent_table_name].cols.values()
                if col.web_visible or col.table_name == 'auto_name'
            ]
            parent_visible_col_names = sorted(
                parent_visible_col_names,
                key=lambda x: parent_table_obj.cols[x].order
            )
            try:
                fk_idx = parent_visible_col_names.index(fk) + 4
            except ValueError:
                print(fk)
                raise ValueError
            col_letter = get_column_letter(fk_idx)

            formula1 = f'=bks_{parent_table_obj.label}!${col_letter}$8:' \
                       f'${col_letter}${str(select_values_rows_count + 9)}'

        if direction == 'vertical':
            cell_col = ws_booking.cell(row=dst_row + col_idx, column=dst_col + 3, value=col_label)
            ws_booking.cell(row=dst_row + col_idx, column=dst_col + 2, value=col_obj.col_name)
            cell_default = ws_booking.cell(row=dst_row + col_idx, column=dst_col + 4, value=col_obj.default)
            if col_obj.foreign_key and not pd.isna(col_obj.foreign_key):
                db_name, foreign_table, foreign_key = col_obj.foreign_key.split('.')
                ws_booking.cell(
                    row=dst_row + col_idx,
                    column=dst_col + 5,
                    value='fk=' + tables[foreign_table].label
                )
            if values:
                if len(values[col_name]) > 0:
                    cell_default.value = values[col_name][0]
        elif direction == 'horizontal':
            cell_col = ws_booking.cell(row=dst_row, column=dst_col + col_idx + 3, value=col_label)
            ws_booking.cell(row=dst_row + 1, column=dst_col + col_idx + 3, value=col_obj.col_name)
            cell_default = ws_booking.cell(row=dst_row + 2, column=dst_col + col_idx + 3, value=col_obj.default)
            if col_obj.foreign_key and not pd.isna(col_obj.foreign_key):
                db_name, foreign_table, foreign_key = col_obj.foreign_key.split('.')
                ws_booking.cell(
                    row=dst_row - 1,
                    column=dst_col + col_idx + 3,
                    value='fk=' + tables[foreign_table].label
                )
            if values:
                try:
                    col_values = values[col_name]

                    for r, value in enumerate(col_values.values()):
                        cell_value = ws_booking.cell(row=dst_row + 2 + r + 1, column=dst_col + col_idx + 3, value=value)
                        if not is_selected_values:
                            ws_booking.cell(row=dst_row + 2 + r + 1, column=dst_col, value=f'row{value_row_index}')
                            apply_cell_format(
                                cell_src=cell_value,
                                cell_target=cell_formats[col_obj.web_obj]
                            )
                            apply_conditional_format(
                                ws_src=ws_booking,
                                cell_src=cell_value,
                                ws_target=ws_cell_format,
                                cell_target=cell_formats[col_obj.web_obj]
                            )

                            if col_obj.foreign_key:
                                apply_data_validation(sheet=ws_booking, cell=cell_value, col_obj=col_obj,
                                                      formula1=formula1)

                            if col_obj.web_obj == 'radio':
                                apply_radio_validation(
                                    sheet=ws_booking, cell=cell_value
                                )

                            value_row_index += 1
                except KeyError:
                    pass

        else:
            raise ValueError

        apply_cell_format(
            cell_src=cell_col,
            cell_target=cell_formats['column'])

        apply_cell_format(
            cell_src=cell_default,
            cell_target=cell_formats[col_obj.web_obj]
        )

        # cell conditional format
        apply_conditional_format(
            ws_src=ws_booking,
            cell_src=cell_default,
            ws_target=ws_cell_format,
            cell_target=cell_formats[col_obj.web_obj]
        )

        # cell validation
        if col_obj.foreign_key and not pd.isna(col_obj.foreign_key):
            apply_data_validation(
                sheet=ws_booking,
                cell=cell_default,
                col_obj=col_obj,
                formula1=formula1
            )

        if col_obj.web_obj == 'radio':
            apply_radio_validation(
                sheet=ws_booking,
                cell=cell_default
            )

    return value_row_index


def render_booking_xl_sheet(output_path, template_path, data_tree: DataTree, con):

    tables = data_tree.tables
    root = data_tree.root
    select_values = data_tree.get_parents_select_values()

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path, keep_vba=True)
    ws_cell_format = wb['cell_format']
    ws_booking = wb.copy_worksheet(wb['bks_root'])
    ws_booking.title = 'bks_' + data_tree.table.label

    data_cell_format = pd.read_excel(template_path, sheet_name='cell_format')
    cell_formats = defaultdict(lambda: ws_cell_format.cell(row=4, column=2))
    d = dict(zip(
        data_cell_format['cell_type'].values,
        [ws_cell_format.cell(row=r + 2, column=2)
         for r in range(len(data_cell_format['cell_type']))]
    ))
    for k, v in d.items():
        cell_formats[k] = v

    dst_row = 4
    dst_col = 1
    value_row_index = 0

    # fill parent tables
    parents_trees = data_tree.get_all_parents_with_full_value()
    for parent_name, parent in parents_trees.items():
        ws_parent_booking = wb.copy_worksheet(wb['bks_root'])
        ws_parent_booking.title = 'bks_' + parent.table.label
        value_row_index = fill_table(
            root=root,
            ws_booking=ws_parent_booking,
            ws_cell_format=ws_cell_format,
            cell_formats=cell_formats,
            select_values=select_values,
            dst_row=5,
            dst_col=1,
            table=parent.table,
            tables=tables,
            direction='horizontal',
            values={parent.reffed: parent.data.reset_index().to_dict()[parent.reffed]},
            is_selected_values=True,
            show_display_name=True
        )

        ws_parent_booking.row_dimensions[6].hidden = True
        ws_parent_booking.row_dimensions[7].hidden = True
        ws_parent_booking.column_dimensions['C'].hidden = True

    # fill root table
    # fill table label
    table = tables[root]
    fill_table(
        root=root,
        ws_booking=ws_booking,
        ws_cell_format=ws_cell_format,
        cell_formats=cell_formats,
        select_values=select_values,
        dst_row=dst_row,
        dst_col=dst_col,
        table=table,
        tables=tables,
        direction='vertical',
        values=data_tree.data.to_dict(orient='list'),
        show_display_name=False
    )

    dst_row = ws_booking.max_row + 3

    # fill child tables
    for child in data_tree.cs:
        value_row_index = fill_table(
            root=root,
            ws_booking=ws_booking,
            ws_cell_format=ws_cell_format,
            cell_formats=cell_formats,
            select_values=select_values,
            dst_row=dst_row,
            dst_col=dst_col,
            table=child.table,
            direction='horizontal',
            tables=tables,
            values=child.data.to_dict(),
            start_value_row_index=value_row_index,
            show_display_name=False
        )

        # hide format row
        ws_booking.row_dimensions[dst_row + 1].hidden = True
        ws_booking.row_dimensions[dst_row + 2].hidden = True

        dst_row = ws_booking.max_row + 3

    ws_booking.column_dimensions['C'].hidden = True

    wb.save(output_path)

