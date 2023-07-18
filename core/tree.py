from typing import Set
import os
from copy import copy, deepcopy

import pandas as pd
from pandas import ExcelWriter
from openpyxl import Workbook

from helper_function.hf_string import to_json_str
from helper_function.excel import fit_col_width
from helper_function.func import *
from core.meta_objs import get_table_objs, JsonObj

from core.curd import CURD


class Tree(JsonObj):
    def __init__(
            self,
            tree=None,
            root=None,
            cst_pki=pd.DataFrame(),
            tables: dict = None,
            ref=None,
            reffed=None
    ):

        if tree:
            self.root = tree.root
            self.cst_pki = copy(tree.cst_pki)
            self.tables = tree.tables
            self.ref = tree.ref
            self.reffed = tree.reffed
            self.table = tree.table
            self.pk = tree.pk

            self.ref_info = tree.cst_pki[
                (tree.cst_pki['CONSTRAINT_NAME'] != 'PRIMARY') &
                (tree.cst_pki['TABLE_NAME'] == tree.root) &
                ~(pd.isna(tree.cst_pki['REFERENCED_TABLE_NAME']))
            ]
            self.reffed_info = tree.cst_pki[
                tree.cst_pki['REFERENCED_TABLE_NAME'] == tree.root
            ]

            self.parents = tree.parents
            self.children = tree.children
            self.node_names = tree.node_names
            self.booking_sequence = tree.booking_sequence
            self.reading_sequence = tree.reading_sequence
            return

        if not root:
            print('root missing')
            raise ValueError

        if len(cst_pki) == 0:
            print(f'missing cols constraint and primary info. tree: {root} ')
            raise ValueError

        if not tables:
            print(f'missing table obj. tree: {root} ')
            raise ValueError

        self.root = root
        self.cst_pki = copy(cst_pki)
        self.ref = ref
        self.reffed = reffed

        self.parents = []
        self.children = []

        JsonObj.__init__(self)

        self.pk = cst_pki[
            (cst_pki['CONSTRAINT_NAME'] == 'PRIMARY') &
            (cst_pki['TABLE_NAME'] == root)
        ]['COLUMN_NAME'].values[0]
        self.tables = tables
        self.table = tables[root]

        self.ref_info = cst_pki[
            (cst_pki['CONSTRAINT_NAME'] != 'PRIMARY') &
            (cst_pki['TABLE_NAME'] == root) &
            ~(pd.isna(cst_pki['REFERENCED_TABLE_NAME']))
        ]
        self.reffed_info = cst_pki[cst_pki['REFERENCED_TABLE_NAME'] == root]

        for i, parents_cst_row in self.ref_info.iterrows():
            parent_root = parents_cst_row['REFERENCED_TABLE_NAME']
            parent_ref = parents_cst_row['COLUMN_NAME']
            parent_reffed = parents_cst_row['REFERENCED_COLUMN_NAME']
            parents_cst_pki = cst_pki[
                (cst_pki['TABLE_NAME'] != self.root) &
                (cst_pki['REFERENCED_TABLE_NAME'] != parent_root)
            ]
            parent = Tree(
                root=parents_cst_row['REFERENCED_TABLE_NAME'],
                cst_pki=parents_cst_pki,
                ref=parent_ref,
                reffed=parent_reffed,
                tables=self.tables
            )
            self.parents.append(parent)

        for i, child_cst_row in self.reffed_info.iterrows():
            child_root = child_cst_row['TABLE_NAME']
            child_ref = child_cst_row['COLUMN_NAME']
            child_reffed = child_cst_row['REFERENCED_COLUMN_NAME']
            children_cst_pki = cst_pki[
                (cst_pki['REFERENCED_TABLE_NAME'] != self.root) &
                (cst_pki['TABLE_NAME'] != self.root)
                ]
            child = Tree(
                root=child_root,
                cst_pki=children_cst_pki,
                ref=child_ref,
                reffed=child_reffed,
                tables=self.tables
            )
            self.children.append(child)

        self.node_names = self.get_node_names()
        self.booking_sequence = self.get_booking_sequence()
        self.reading_sequence = self.get_reading_sequence()

    def __bool__(self):
        if self.root:
            return True
        else:
            return False

    def __gt__(self, other):
        return self.root in [p for p in other.all_parenthood_names()]

    def __lt__(self, other):
        return self.root in [c for c in other.all_childhood_names()]

    def __ge__(self, other):
        return self.root in [p for p in other.all_parenthood_names()] or self.root == other.root

    def __le__(self, other):
        return self.root in [c for c in other.all_childhood_names()] or self.root == other.root

    def __eq__(self, other):
        return self.root == other.root

    def all_parenthood_names(self):
        res = set([p.root for p in self.parents])
        for p in self.parents:
            res |= p.all_parenthood_names()
        for c in self.children:
            res |= c.all_parenthood_names()
        return res

    def all_childhood_names(self):
        res = set([c.root for c in self.children])
        for c in self.children:
            res |= c.all_parenthood_names()
        return res

    def json_obj_base(self, with_table=False):
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'pk', 'ref', 'reffed', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        for parent in self.parents:
            json_obj['parents'].append(parent.json_obj_base(with_table=with_table))
        for child in self.children:
            json_obj['children'].append(child.json_obj_base(with_table=with_table))

        json_obj['children'].sort(key=lambda x: x['table']['index'])
        json_obj['table'] = self.table.to_json_obj(exclude_attrs=['fields_info'])

        cols = self.table.cols
        col_orders = dict(zip(cols.keys(), [col.order for col in cols.values()]))

        p_orders = {parent.ref: cols[parent.ref].order for parent in self.parents}
        sorted_orders = sorted([i for i in p_orders.values()])
        p_orders = {key: sorted_orders.index(value) for key, value in p_orders.items()}

        col_orders = {key: value for key, value in col_orders.items()
                      if key not in p_orders.keys()}

        sorted_orders = sorted([i for i in col_orders.values()])
        col_orders = {key: sorted_orders.index(value) + len(p_orders) for key, value in col_orders.items()}

        for key, order in p_orders.items():
            json_obj['table']['cols'][key]['order'] = order

        for key, order in col_orders.items():
            json_obj['table']['cols'][key]['order'] = order

        return json_obj

    @property
    def json_obj(self):
        return self.json_obj_base(with_table=True)

    @property
    def json_obj_without_table(self):
        return self.json_obj_base(with_table=False)
    
    def p_(self, p_name, ref=None):
        if not ref:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name
                ), None
            )
        else:
            return next(
                (
                    parent for parent in self.parents
                    if parent.root == p_name and parent.ref == ref
                ), None
            )

    def c_(self, c_name, reffed=None):
        if not reffed:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name
                ), None
            )
        else:
            return next(
                (
                    child for child in self.children
                    if child.root == c_name and child.reffed == reffed
                ), None
            )
    
    @property
    def p(self):
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        return PropertyIndexer(self, 'c_')
    
    def get_node_names(self):
        res = {self.root}
        for parent in self.parents:
            res |= parent.node_names
        for child in self.children:
            res |= child.node_names
        return res

    def get_booking_sequence(self, cst_pki=pd.DataFrame()):
        if len(self.node_names) == 0:
            return []
        elif len(self.node_names) == 1:
            return list(self.node_names)
        else:
            if len(cst_pki) == 0:
                cst_pki = copy(self.cst_pki)

            cst_pki = cst_pki[
                (cst_pki['CONSTRAINT_NAME'] != 'PRIMARY') &
                (~pd.isna(cst_pki['REFERENCED_TABLE_NAME']))
            ]

            booking_seq = []
            while len(cst_pki) > 0:
                ref_tables = set(cst_pki['TABLE_NAME'].to_list())
                reffed_tables = set(cst_pki['REFERENCED_TABLE_NAME'].to_list())

                booking_seq.extend(sorted(list(reffed_tables - ref_tables)))

                cst_pki = cst_pki[~cst_pki['REFERENCED_TABLE_NAME'].isin(booking_seq)]

            booking_seq = booking_seq + list(set(self.node_names) - set(booking_seq))

            booking_seq = sorted(
                self.node_names,
                key=lambda x: booking_seq.index(x)
            )

            return booking_seq

    def get_reading_sequence(self):
        booking_seq = self.get_booking_sequence()
        root_table_index = booking_seq.index(self.root)
        reading_seq = booking_seq[root_table_index:] + list(booking_seq[:root_table_index].__reversed__())
        return reading_seq


class DataTree(Tree):
    def __init__(
            self,
            tree=None,
            root=None,
            cst_pki=pd.DataFrame(),
            tables: dict = None,
            ref=None,
            reffed=None,
            relevant_data_set=None,
    ):
        super().__init__(
            tree=tree,
            root=root,
            cst_pki=cst_pki,
            tables=tables,
            ref=ref,
            reffed=reffed,
        )

        if relevant_data_set is None:
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        try:
            self.data = relevant_data_set[self.root]
        except KeyError:
            self.data = pd.DataFrame(columns=self.table.cols.keys())

        self.relevant_data_set = relevant_data_set

    def __len__(self):
        return len(self.data)

    def __getitem__(self, key):
        return self.data[key]

    @property
    def loc(self):
        return self.data.loc

    @property
    def iloc(self):
        return self.data.iloc

    def iterrows(self):
        return self.data.iterrows()

    def itertrees(self):
        for i, row in self.iterrows():
            relevant_data_set = deepcopy(self.relevant_data_set)
            relevant_data_set[self.root] = row.to_frame().transpose()
            datatree = DataTree(
                tree=self,
                relevant_data_set={}
            )
            datatree.from_relevant_data_set(relevant_data_set)
            yield i, datatree

    @staticmethod
    def update_values_map(values_map, tree, root_data):
        def update_values_map_sub(sub_tree, from_key, to_key):
            if len(root_data) == 0:
                adding_data = set()
            else:
                adding_data = set(root_data[from_key].dropna())
            try:
                values_map[sub_tree.root][to_key]['values'] = \
                    values_map[sub_tree.root][to_key]['values'] | adding_data
            except KeyError:
                values_map[sub_tree.root] = {}
                values_map[sub_tree.root][to_key] = {
                    'tree': sub_tree,
                    'values': adding_data
                }

        for p in tree.parents:
            update_values_map_sub(sub_tree=p, from_key=p.ref, to_key=p.reffed)

        for c in tree.children:
            update_values_map_sub(sub_tree=c, from_key=c.reffed, to_key=c.ref)

        return values_map

    def from_sql(
            self,
            index_field: str = None,
            index_values: Set[str] = (),
            con=None
    ):
        # root data
        if index_field is None:
            where_str = ''
        else:
            values_str = ', '.join(['"%s"' % index_value for index_value in index_values])
            where_str = f'WHERE `{index_field}` in ({values_str})'

        sql = f'SELECT * FROM {self.root} {where_str}'
        root_data = pd.read_sql(sql=sql, con=con, index_col=self.pk)
        if len(root_data) == 0:
            print(f'empty result for root: "{self.root}" \n'
                  f'with index field: "{index_field}" \n'
                  f'with values in {index_values}')
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        else:
            relevant_data_set = {self.root: root_data}

            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # relevant data
            reading_seq = self.reading_sequence[1:]

            for node_name in reading_seq:
                where_str_list = []
                for field, tree_values in values_map[node_name].items():
                    values = tree_values['values']
                    values_str = ', '.join(['"%s"' % value for value in values])
                    where_str = f'(`{field}` in ({values_str}))'
                    where_str_list.append(where_str)

                all_where_str = ' or '.join(where_str_list)

                sql = f'SELECT * FROM {node_name} WHERE {all_where_str}'
                data = pd.read_sql(sql=sql, con=con, index_col=self.tables[node_name].pk)
                relevant_data_set[node_name] = data

                for field, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

        self.data = relevant_data_set[self.root]
        self.relevant_data_set = relevant_data_set

    def from_relevant_data_set(self, relevant_data_set=None):

        if relevant_data_set is None:
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))

        root_data = relevant_data_set[self.root]
        if len(root_data) == 0:
            relevant_data_set = dict(zip(
                self.node_names,
                [pd.DataFrame(columns=self.tables[node_name].cols.keys())
                 for node_name in self.node_names]
            ))
        else:
            values_map = self.update_values_map(
                values_map={},
                tree=self,
                root_data=root_data
            )

            # relevant data
            reading_seq = self.reading_sequence[1:]

            for node_name in reading_seq:
                where_str_list = []
                data = pd.DataFrame(columns=self.table.cols.keys())
                for field, tree_values in values_map[node_name].items():
                    values = tree_values['values']
                    values_str = ', '.join(['"%s"' % value for value in values])
                    data = relevant_data_set[node_name]
                    where_str = f'(data["{field}"].isin([{values_str}]))'
                    where_str_list.append(where_str)

                all_where_str = ' | '.join(where_str_list)

                relevant_data_set[node_name] = eval(f'data[{all_where_str}]')

                for field, tree_values in values_map[node_name].items():
                    values_map = self.update_values_map(
                        values_map=values_map,
                        tree=tree_values['tree'],
                        root_data=data
                    )

        self.data = root_data
        self.relevant_data_set = relevant_data_set

    @check_param_valid_range(
        [
            'if_data_exists'
        ],
        [
            ['fail', 'skip', 'merge', 'update']
        ]
    )
    def to_sql(self, curd, if_data_exists='fail'):
        stmt = curd.stmt_df_merge(
            datatree=self,
            if_data_exists=if_data_exists,
        )
        print(stmt)
        # con.cursor.execute(stmt)

    def p_(self, p_name, ref=None):
        p = DataTree(
            tree=Tree.p_(self, p_name=p_name, ref=ref)
        )
        p.from_relevant_data_set(self.relevant_data_set)
        yield p

    def c_(self, c_name, reffed=None):
        c = DataTree(
            tree=Tree.c_(self, c_name=c_name, reffed=reffed)
        )
        c.from_relevant_data_set(self.relevant_data_set)
        return c

    @property
    def p(self):
        return PropertyIndexer(self, 'p_')

    @property
    def c(self):
        return PropertyIndexer(self, 'c_')

    @property
    def ps(self):
        for parent in self.parents:
            p = DataTree(
                tree=parent
            )
            p.from_relevant_data_set(self.relevant_data_set)
            yield p

    @property
    def cs(self):
        for child in self.children:
            c = DataTree(
                tree=child
            )
            c.from_relevant_data_set(self.relevant_data_set)
            yield c

    def json_obj_base(self, with_table=False, with_value=False):
        json_obj = self.to_json_obj_raw(
            include_attrs=['root', 'pk', 'parents', 'children']
        )
        json_obj['parents'] = []
        json_obj['children'] = []
        for parent in self.ps:
            json_obj['parents'].append(parent.json_obj_base(with_table=with_table, with_value=with_value))
        for child in self.cs:
            json_obj['children'].append(child.json_obj_base(with_table=with_table, with_value=with_value))

        if with_table:
            json_obj['table'] = self.table.to_json_obj(include_attrs='all')

        if with_value:
            json_obj['values'] = self.data.to_dict(orient='index')
        return json_obj

    @property
    def json_obj(self):
        return self.json_obj_base(with_table=True, with_value=True)

    @property
    def json_obj_table_only(self):
        return self.json_obj_base(with_table=True, with_value=False)

    @property
    def json_obj_value_only(self):
        return self.json_obj_base(with_table=False, with_value=True)

    def get_parents_select_values(self, con, res=None):
        if not res:
            res = {}
        for p in self.ps:
            if p.root not in list(res.keys()):
                p.from_sql(con=con)
                res[p.root] = p.data[p.reffed].to_list()
                res.update(p.get_parents_select_values(con=con, res=res))
        for c in self.cs:
            res.update(c.get_parents_select_values(con=con, res=res))
        return res

    def to_excel_sheets(self, pth, node_names=None, web_visible_only=False, web_label=False, table_web_label=False):
        writer = ExcelWriter(pth, mode='w', engine='openpyxl', options={'strings_to_urls': False})
        if not node_names:
            node_names = self.node_names
        flag = False

        for node_name in node_names:
            table = self.tables[node_name]
            if web_visible_only:
                columns = [
                    col.field
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            else:
                columns = [
                    col.field
                    for col in table.cols.values()
                    if col.web_visible == 1
                ]

            try:
                df = self.relevant_data_set[node_name][columns]
            except KeyError:
                df = pd.DataFrame(columns=columns)

            if web_label:
                column_mapper = dict(
                    [
                        (col.field, col.web_label)
                        for col in table.cols.values()]
                )
                df.rename(columns=column_mapper, inplace=True)
            if table_web_label:
                node_name = table.comment

            df.to_excel(writer, sheet_name=node_name, index=False,
                        columns=df.columns)
            fit_col_width(writer=writer, df=df, sheet_name=node_name)
            flag = True

        if flag:
            writer.save()
            writer.close()
        else:
            del writer
            os.remove(pth)

    def to_excel_detail_sheet(self, pth):
        child_data = {}
        for child in self.children:
            child_data[child.root] = self[child.root].to_dict()
        wb = Workbook()
        ws = wb.worksheets[0]
        pin = 1
        for r, (key, value) in enumerate(self.data.items()):
            if self.table.cols[key].web_visible == 1:
                ws.cell(row=pin, column=1).value = self.table.cols[key].web_label
                ws.cell(row=pin, column=2).value = value
                pin += 1

        pin += 1

        for child_name, child_values in child_data.items():
            ws.cell(row=pin, column=1).value = self.tables[child_name].comment
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    ws.cell(row=pin, column=c_index).value = self.tables[child_name].cols[col].web_label
                    c_index += 1
            pin += 1
            c_index = 1
            for c, (col, value) in enumerate(child_values.items()):
                if self.tables[child_name].cols[col].web_visible == 1:
                    for r, (index, v) in enumerate(value.items()):
                        ws.cell(row=pin + r, column=c_index).value = v
                    c_index += 1
            pin += len(child_values) + 1

        wb.save(pth)


class Test:
    from sys_init import DB_URL
    curd = CURD(url=DB_URL)
    cst_pki = curd.get_cst_pki()
    tables = get_table_objs(curd=curd)
    t = Tree(root='project', cst_pki=cst_pki, tables=tables)
    dt = DataTree(tree=t)

    def __init__(self):
        pass

    # @print_output
    @timecost(30)
    def tree(self):
        return Tree(root='project', cst_pki=self.cst_pki, tables=self.tables)

    # @print_output
    @timecost(10)
    def tree_json_obj(self):
        jo = self.t.json_obj
        return to_json_str(jo)

    # @print_output
    @timecost(10)
    def tree_json_obj_without_table(self):
        jo_wt = self.t.json_obj_without_table
        return to_json_str(jo_wt)

    # @print_output
    @timecost(10)
    def tree_p(self):
        p = self.t.p['bd']
        return to_json_str(p.json_obj)

    # @print_output
    @timecost(10)
    def tree_c(self):
        p = self.t.c['project_inst']
        return to_json_str(p.json_obj)

    # @print_output
    @timecost(10)
    def tree_get_node_names(self):
        return self.t.get_node_names()

    # @print_output
    @timecost(10)
    def tree_get_booking_seq(self):
        return self.t.get_booking_sequence()

    @timecost(10)
    def datatree(self):
        return DataTree(
            tree=Tree(root='project', cst_pki=self.cst_pki, tables=self.tables)
        )

    @timecost(10)
    def datatree_with_exist_tree(self):
        return DataTree(tree=self.t)

    @timecost(1)
    def datatree_from_sql(self):
        con = self.curd.engine.connect()
        self.dt.from_sql(
            index_field='nick',
            index_values={'华泰南方1期'},
            con=con
        )
        con.close()

    @timecost(1)
    def datatree_from_sql_all(self):
        con = self.curd.engine.connect()
        self.dt.from_sql(
            con=con
        )
        con.close()

    @timecost(1)
    def datatree_itertrees(self):
        names = []
        for i, datat in self.dt.itertrees():
            names.append(datat['name'])

    @timecost(10)
    def datatree_loop_p(self):
        for _ in self.dt.ps:
            pass

    @timecost(10)
    def datatree_loop_c(self):
        for _ in self.dt.cs:
            pass

    @timecost(10)
    def datatree_json_obj_base(self):
        return self.dt.json_obj_base()

    @timecost(10)
    def datatree_json_obj(self):
        return self.dt.json_obj

    @timecost(10)
    def datatree_json_obj_table_only(self):
        return self.dt.json_obj_table_only

    @timecost(10)
    def datatree_json_obj_value_only(self):
        return self.dt.json_obj_value_only

    @timecost(10)
    def datatree_from_relevant_data_set(self):
        return self.dt.from_relevant_data_set(relevant_data_set=self.dt.relevant_data_set)
    
    def datatree_fancy_indexing(self):
        self.datatree_from_sql()
        return self.dt.c['project_level', 'name'].data
    
    def _datatree_command_test(self):
        print('command_test:')
        self.datatree_from_sql()
        dt = self.dt
        while True:
            s = input('commands:')
            try:
                print(eval(s))
            except Exception as e:
                print(e)

    def all_tests(self, prefix=''):
        for func_name in self.__dir__():
            if func_name[:len(prefix)] == prefix:
                exec(f'self.{func_name}()')
    
    
if __name__ == '__main__':
    tt = Test()
    tt.all_tests('tree')
